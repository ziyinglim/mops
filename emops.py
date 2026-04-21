"""
EMOPS Scraper — Company Profile + Balance Sheet
Reads from: https://emops.twse.com.tw
Output: terminal preview + Excel file + JSON archive
"""

import asyncio
import hashlib
import json
import logging
import re
import warnings
from datetime import datetime, timezone
from pathlib import Path

import httpx
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", message="Unverified HTTPS request")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("emops")

# ── Config ────────────────────────────────────────────────────────────────────

EMOPS_HOST = "https://emops.twse.com.tw"
TYPEK_OPTIONS = ["sii", "otc", "rotc", "co"]

OUTPUT_DIR = Path("output")
ARCHIVE_DIR = Path("storage/archive")
STATE_DIR = Path("storage/state")

WATCHLIST = [
    {"stock_code": "2881", "name_en": "Fubon Financial Holding"},
    {"stock_code": "2882", "name_en": "Cathay Life Insurance"},
    {"stock_code": "2891", "name_en": "CBC Financial Holding"},
    {"stock_code": "2330", "name_en": "Taiwan Semiconductor Manufacturing Company"},
    {"stock_code": "5874", "name_en": "Nan Shan Life Insurance"},
    {"stock_code": "2317", "name_en": "Foxconn Technology Group"},
    {"stock_code": "2886", "name_en": "Mega International Commercial Bank"},
    {"stock_code": "2880", "name_en": "Hua Nan Commercial Bank"},
    {"stock_code": "5857", "name_en": "Land Bank of Taiwan"},
    {"stock_code": "2888", "name_en": "Shin Kong Life Insurance"},
    {"stock_code": "2801", "name_en": "Chang Hwa Bank"},
    {"stock_code": "2890", "name_en": "Bank SinoPac"},
    {"stock_code": "5876", "name_en": "Shanghai Commercial & Savings Bank"},
    {"stock_code": "2885", "name_en": "Yuanta Commercial Bank"},
    {"stock_code": "2833", "name_en": "Taiwan Life Insurance"},
    {"stock_code": "2867", "name_en": "Mercuries Life Insurance"},
    {"stock_code": "5873", "name_en": "TransGlobe Life Insurance"},
    {"stock_code": "2382", "name_en": "Quanta Computer"},
    {"stock_code": "3231", "name_en": "Wistron Corporation"},
    {"stock_code": "3711", "name_en": "ASE Technology Holding"},
    {"stock_code": "5859", "name_en": "Farglory Life Insurance"},
    {"stock_code": "2454", "name_en": "MediaTek"},
    {"stock_code": "2897", "name_en": "O-Bank"},
    {"stock_code": "2002", "name_en": "China Steel Corporation"},
]

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": f"{EMOPS_HOST}/server-java/t58query",
    "Origin": EMOPS_HOST,
    "Content-Type": "application/x-www-form-urlencoded",
}

# ── HTTP Client ───────────────────────────────────────────────────────────────

async def post_emops(path: str, stock_code: str, extra: dict = None) -> str | None:
    url = EMOPS_HOST + path
    base_data = {"co_id": stock_code, **(extra or {})}

    async with httpx.AsyncClient(headers=_HEADERS, timeout=30, follow_redirects=True, verify=False) as client:
        try:
            await client.get(f"{EMOPS_HOST}/server-java/t58query")
        except Exception:
            pass

        for typek in TYPEK_OPTIONS:
            data = {**base_data, "TYPEK": typek}
            try:
                resp = await client.post(url, data=data)
                resp.raise_for_status()
                html = resp.text
                if html and len(html) > 200 and "error" not in html[:200].lower():
                    logger.info("POST %s [%s] TYPEK=%s → %d chars", path, stock_code, typek, len(html))
                    return html
            except Exception as exc:
                logger.warning("POST failed [%s] TYPEK=%s: %s", stock_code, typek, exc)

    logger.error("All TYPEK attempts failed for %s %s", path, stock_code)
    return None

# ── Profile Scraper ───────────────────────────────────────────────────────────

async def scrape_profile(stock_code: str) -> dict:
    html = await post_emops("/server-java/t146sb05_e", stock_code, {"step": "0"})
    if not html:
        return {"stock_code": stock_code, "error": "No response"}

    soup = BeautifulSoup(html, "lxml")
    return {
        "stock_code": stock_code,
        "company_name_en": _find_column_field(soup, ["Company Name"]),
        "chairman":        _find_column_field(soup, ["Chairman"]),
        "general_manager": _find_column_field(soup, ["General Manager"]),
        "telephone":       _find_column_field(soup, ["Telephone"]),
        "web_address":     _find_column_field(soup, ["Web Address"]),
        "address":         _find_address(soup),
    }

def _find_column_field(soup: BeautifulSoup, labels: list[str]) -> str:
    for label in labels:
        for cell in soup.find_all(["td", "th"], string=lambda t: t and label.lower() in t.lower()):
            row = cell.find_parent("tr")
            if not row:
                continue
            siblings = row.find_all(["td", "th"])
            try:
                col_idx = siblings.index(cell)
            except ValueError:
                continue
            next_row = row.find_next_sibling("tr")
            if next_row:
                next_cells = next_row.find_all(["td", "th"])
                if col_idx < len(next_cells):
                    val = next_cells[col_idx].get_text(strip=True)
                    if val and val.lower() != label.lower():
                        return val
    return ""

def _find_address(soup: BeautifulSoup) -> str:
    for td in soup.find_all("td", attrs={"colspan": True}):
        text = td.get_text(strip=True)
        if text and any(kw in text for kw in ["Road", "St.", "Ave", "F.,", "No.", "路", "街", "號"]):
            return _dedup_address(text)
    return ""

def _dedup_address(text: str) -> str:
    """Remove duplicated trailing segments e.g. 'TaiwanTaipei Taiwan' → 'Taiwan'."""
    # Split on known country/city duplications
    for marker in ["Taiwan", "R.O.C"]:
        idx = text.find(marker)
        if idx != -1:
            return text[:idx + len(marker)].strip(" ,")
    return text

# ── Balance Sheet Scraper ─────────────────────────────────────────────────────

async def scrape_balance_sheet(stock_code: str) -> dict:
    html = await post_emops("/server-java/t164sb03_e", stock_code, {"step": "current"})
    if not html:
        return {"stock_code": stock_code, "error": "No response"}

    soup = BeautifulSoup(html, "lxml")

    period = _extract_period(soup)

    currency = "TWD (thousands)" if "千元" in soup.get_text() else "TWD"

    total_assets        = _find_balance_value(soup, ["Total assets", "Total Assets", "資產總計"])
    investment_property = _find_balance_value(soup, ["Investment property, net", "Investment property", "投資性不動產淨額", "投資性不動產"])

    return {
        "stock_code": stock_code,
        "period": period,
        "currency": currency,
        "total_assets_raw": total_assets,
        "total_assets_numeric": _parse_number(total_assets),
        "investment_property_raw": investment_property,
        "investment_property_numeric": _parse_number(investment_property),
    }

def _extract_period(soup: BeautifulSoup) -> str:
    """Extract reporting date from balance sheet header, return YYYY/MM/DD."""
    full_text = soup.get_text(" ")

    # Pattern: 2025/12/31 or 2025-12-31
    m = re.search(r"(20\d{2})[/-](\d{2})[/-](\d{2})", full_text)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"

    # ROC calendar: 民國114年12月31日 → convert to 2025/12/31
    m = re.search(r"民國\s*(\d{2,3})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", full_text)
    if m:
        year = int(m.group(1)) + 1911
        return f"{year}/{int(m.group(2)):02d}/{int(m.group(3)):02d}"

    # English month: December 31, 2025
    months = {"January":"01","February":"02","March":"03","April":"04","May":"05","June":"06",
              "July":"07","August":"08","September":"09","October":"10","November":"11","December":"12"}
    m = re.search(r"(January|February|March|April|May|June|July|August|September|October|November|December)"
                  r"\s+(\d{1,2}),?\s*(20\d{2})", full_text)
    if m:
        return f"{m.group(3)}/{months[m.group(1)]}/{int(m.group(2)):02d}"

    return ""


def _find_balance_value(soup: BeautifulSoup, labels: list[str]) -> str:
    for label in labels:
        for cell in soup.find_all("td", string=lambda t: t and label.lower() in t.replace("\u3000", "").strip().lower()):
            value_td = cell.find_next_sibling("td")
            if value_td:
                val = value_td.get_text(strip=True)
                if val:
                    return val
    return ""

def _parse_number(raw: str) -> float | None:
    if not raw:
        return None
    cleaned = raw.replace(",", "").replace(" ", "")
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = "-" + cleaned[1:-1]
    try:
        return float(cleaned)
    except ValueError:
        return None

# ── Change Detection ──────────────────────────────────────────────────────────

def detect_changes(records: list[dict], state_path: Path, key_fields: list[str]) -> list[dict]:
    stored = _load_json(state_path) or {}
    updated = dict(stored)
    now = datetime.now(timezone.utc).isoformat()

    for record in records:
        key = "|".join(str(record.get(f, "")) for f in key_fields)
        current_hash = hashlib.sha256(
            json.dumps({k: v for k, v in record.items() if k not in {"scraped_at", "status", "hash"}},
                       sort_keys=True, ensure_ascii=False).encode()
        ).hexdigest()
        record["hash"] = current_hash
        record["scraped_at"] = now

        if key not in stored:
            record["status"] = "NEW"
        elif stored[key]["hash"] != current_hash:
            record["status"] = "CHANGED"
        else:
            record["status"] = "UNCHANGED"

        updated[key] = {"hash": current_hash, "last_seen": now}

    state_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.write_text(json.dumps(updated, indent=2, ensure_ascii=False), encoding="utf-8")
    return records

# ── Archive ───────────────────────────────────────────────────────────────────

def archive(stock_code: str, category: str, records: list[dict]):
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    path = ARCHIVE_DIR / f"{stock_code}_{category}_{ts}.json"
    path.write_text(json.dumps({"stock_code": stock_code, "category": category,
                                "records": records}, indent=2, ensure_ascii=False), encoding="utf-8")

def _load_json(path: Path) -> dict | None:
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return None

# ── Excel Output ──────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="1F3864")
_NEW_FILL  = PatternFill("solid", fgColor="C6EFCE")
_CHG_FILL  = PatternFill("solid", fgColor="FFEB9C")
_HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_BODY_FONT = Font(name="Calibri", size=10)

def write_excel(profiles: list[dict], balance_sheets: list[dict]) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    path = OUTPUT_DIR / f"EMOPS_{ts}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Summary sheet
    ws = wb.create_sheet("Summary")
    _header(ws, ["Run Date", "Companies", "Profiles OK", "Balance Sheets OK"])
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        len(WATCHLIST),
        sum(1 for p in profiles if not p.get("error")),
        sum(1 for b in balance_sheets if not b.get("error")),
    ])
    _autofit(ws)

    # Company Profile sheet
    ws = wb.create_sheet("CompanyProfile")
    _header(ws, ["Stock Code", "Company Name (EN)", "Chairman", "General Manager",
                 "Telephone", "Web Address", "Address", "Status", "Scraped At"])
    for i, r in enumerate(profiles, 2):
        ws.append([r.get("stock_code"), r.get("company_name_en"), r.get("chairman"),
                   r.get("general_manager"), r.get("telephone"), r.get("web_address"),
                   r.get("address"), r.get("status"), r.get("scraped_at")])
        _status_fill(ws, i, r.get("status"))
    _autofit(ws)

    # Balance Sheet sheet
    ws = wb.create_sheet("BalanceSheet")
    _header(ws, ["Stock Code", "Period", "Currency",
                 "Total Assets (Raw)", "Total Assets (Numeric)",
                 "Investment Property (Raw)", "Investment Property (Numeric)",
                 "AUM (bn)", "Status", "Scraped At"])
    for i, r in enumerate(balance_sheets, 2):
        aum = f"=IF(E{i}<>\"\",E{i}/1000000000,\"\")"
        ws.append([r.get("stock_code"), r.get("period"), r.get("currency"),
                   r.get("total_assets_raw"), r.get("total_assets_numeric"),
                   r.get("investment_property_raw"), r.get("investment_property_numeric"),
                   aum, r.get("status"), r.get("scraped_at")])
        _status_fill(ws, i, r.get("status"))
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=7):
        for cell in row:
            cell.number_format = "#,##0"
    _autofit(ws)

    wb.save(path)
    logger.info("Excel saved: %s", path)
    return path

def _header(ws, cols):
    for c, h in enumerate(cols, 1):
        cell = ws.cell(1, c, h)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _status_fill(ws, row, status):
    if status == "NEW":
        for cell in ws[row]: cell.fill = _NEW_FILL
    elif status == "CHANGED":
        for cell in ws[row]: cell.fill = _CHG_FILL

def _autofit(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 50)

# ── Terminal Preview ──────────────────────────────────────────────────────────

def print_results(profiles: list[dict], balance_sheets: list[dict]):
    bs_map = {b["stock_code"]: b for b in balance_sheets}

    print("\n" + "═" * 80)
    print("  EMOPS RESULTS")
    print("═" * 80)
    for p in profiles:
        code = p.get("stock_code")
        b = bs_map.get(code, {})
        print(f"\n  [{code}] {p.get('company_name_en', 'N/A')}")
        if p.get("error"):
            print(f"    Profile ERROR     : {p['error']}")
        else:
            print(f"    Address           : {p.get('address', 'N/A')}")
            print(f"    Telephone         : {p.get('telephone', 'N/A')}")
            print(f"    Website           : {p.get('web_address', 'N/A')}")
        if b.get("error"):
            print(f"    Balance Sht ERROR : {b['error']}")
        else:
            print(f"    Period            : {b.get('period', 'N/A')}")
            print(f"    Currency          : {b.get('currency', 'N/A')}")
            print(f"    Total Assets      : {b.get('total_assets_raw', 'N/A')}")
            print(f"    Investment Prop   : {b.get('investment_property_raw', 'N/A')}")
    print("\n" + "═" * 80 + "\n")

# ── Main ──────────────────────────────────────────────────────────────────────

async def run(companies: list[str] | None = None, export_excel: bool = True):
    watchlist = WATCHLIST if not companies else [w for w in WATCHLIST if w["stock_code"] in companies]
    logger.info("Running EMOPS for %d companies", len(watchlist))

    profiles, balance_sheets = [], []

    for entry in watchlist:
        code = entry["stock_code"]
        logger.info("── %s %s", code, entry["name_en"])

        profile = await scrape_profile(code)
        profile = detect_changes([profile], STATE_DIR / f"{code}_profile.json", ["stock_code"])[0]
        profiles.append(profile)
        archive(code, "profile", [profile])

        bs = await scrape_balance_sheet(code)
        bs = detect_changes([bs], STATE_DIR / f"{code}_balance_sheet.json", ["stock_code"])[0]
        balance_sheets.append(bs)
        archive(code, "balance_sheet", [bs])

    print_results(profiles, balance_sheets)

    if export_excel:
        write_excel(profiles, balance_sheets)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="EMOPS Scraper")
    parser.add_argument("--companies", nargs="+", help="Limit to specific stock codes e.g. 2882 2330")
    parser.add_argument("--no-excel", action="store_true", help="Print results only, skip Excel export")
    args = parser.parse_args()

    asyncio.run(run(companies=args.companies, export_excel=not args.no_excel))
