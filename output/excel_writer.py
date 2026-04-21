"""
Writes extraction results to a multi-sheet Excel workbook using openpyxl.
Sheets: Summary, CompanyProfile, BalanceSheet, FundCommitments, PeopleMoves, JIRA_Upload
Financial calculations use live Excel formulas, not hardcoded values.
"""
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # Dark navy
_SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")  # Blue
_NEW_FILL = PatternFill("solid", fgColor="C6EFCE")      # Green
_CHANGED_FILL = PatternFill("solid", fgColor="FFEB9C")  # Yellow
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_BODY_FONT = Font(name="Calibri", size=10)


def write_excel(
    output_path: Path,
    profiles: list[dict],
    balance_sheets: list[dict],
    fund_commitments: list[dict],
    people_moves: list[dict],
    watchlist: list[dict],
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    _write_summary(wb, profiles, balance_sheets, fund_commitments, people_moves, watchlist)
    _write_company_profile(wb, profiles)
    _write_balance_sheet(wb, balance_sheets)
    _write_fund_commitments(wb, fund_commitments)
    _write_people_moves(wb, people_moves)
    _write_jira_upload(wb, fund_commitments, people_moves, watchlist)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Excel written to %s", output_path)
    return output_path


# ── Summary Sheet ─────────────────────────────────────────────────────────────

def _write_summary(wb, profiles, balance_sheets, fund_commitments, people_moves, watchlist):
    ws = wb.create_sheet("Summary")
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    _header_row(ws, 1, [
        "Run Date", "Companies Tracked", "New Fund Commitments",
        "Changed Fund Commitments", "New People Moves", "Changed People Moves",
        "Balance Sheets Fetched", "Profiles Fetched",
    ])

    new_funds = sum(1 for r in fund_commitments if r.get("status") == "NEW")
    chg_funds = sum(1 for r in fund_commitments if r.get("status") == "CHANGED")
    new_people = sum(1 for r in people_moves if r.get("status") == "NEW")
    chg_people = sum(1 for r in people_moves if r.get("status") == "CHANGED")

    ws.append([now, len(watchlist), new_funds, chg_funds, new_people, chg_people,
               len(balance_sheets), len(profiles)])
    _apply_body_style(ws, 2)
    _autofit(ws)


# ── Company Profile Sheet ─────────────────────────────────────────────────────

def _write_company_profile(wb, profiles):
    ws = wb.create_sheet("CompanyProfile")
    cols = ["Stock Code", "Company Name (ZH)", "Company Name (EN)", "Address", "Telephone", "Web Address", "Status"]
    _header_row(ws, 1, cols)

    for i, r in enumerate(profiles, start=2):
        ws.append([
            r.get("stock_code"), r.get("company_name_zh"), r.get("company_name_en"),
            r.get("address"), r.get("telephone"), r.get("web_address"), r.get("status"),
        ])
        _apply_status_fill(ws, i, r.get("status"))

    _autofit(ws)


# ── Balance Sheet Sheet ───────────────────────────────────────────────────────

def _write_balance_sheet(wb, balance_sheets):
    ws = wb.create_sheet("BalanceSheet")
    cols = [
        "Stock Code", "Period", "Currency",
        "Total Assets (Raw)", "Total Assets (Numeric)",
        "Investment Property (Raw)", "Investment Property (Numeric)",
        "AUM (bn)", "Status", "Scraped At",
    ]
    _header_row(ws, 1, cols)

    for i, r in enumerate(balance_sheets, start=2):
        total_col = get_column_letter(5)  # E = Total Assets Numeric
        aum_formula = f"=IF({total_col}{i}<>\"\",{total_col}{i}/1000000000,\"\")"
        ws.append([
            r.get("stock_code"), r.get("period"), r.get("currency"),
            r.get("total_assets_raw"), r.get("total_assets_numeric"),
            r.get("investment_property_raw"), r.get("investment_property_numeric"),
            aum_formula,
            r.get("status"), r.get("scraped_at"),
        ])
        _apply_status_fill(ws, i, r.get("status"))

    # Format numeric columns
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=7):
        for cell in row:
            cell.number_format = "#,##0"
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = "#,##0.0"

    _autofit(ws)


# ── Fund Commitments Sheet ────────────────────────────────────────────────────

def _write_fund_commitments(wb, fund_commitments):
    ws = wb.create_sheet("FundCommitments")
    cols = [
        "Stock Code", "Fund Name", "Fund Type",
        "Commitment Date", "Commitment Amount (Raw)",
        "Commitment Amount (Numeric)", "Currency",
        "Status", "Announcement URL", "Scraped At",
    ]
    _header_row(ws, 1, cols)

    for i, r in enumerate(fund_commitments, start=2):
        ws.append([
            r.get("stock_code"), r.get("fund_name"), r.get("fund_type"),
            r.get("commitment_date"), r.get("commitment_amount_raw"),
            r.get("commitment_amount_numeric"), r.get("commitment_currency"),
            r.get("status"), r.get("announcement_url"), r.get("scraped_at"),
        ])
        _apply_status_fill(ws, i, r.get("status"))

    _autofit(ws)


# ── People Moves Sheet ────────────────────────────────────────────────────────

def _write_people_moves(wb, people_moves):
    ws = wb.create_sheet("PeopleMoves")
    cols = [
        "Stock Code", "Role Type", "New Holder", "Previous Holder",
        "Change Type", "Change Date", "Effective Date",
        "Reason", "Narrative (EN)", "Status", "Announcement URL", "Scraped At",
    ]
    _header_row(ws, 1, cols)

    for i, r in enumerate(people_moves, start=2):
        ws.append([
            r.get("stock_code"), r.get("role_type"), r.get("new_holder"),
            r.get("previous_holder"), r.get("change_type"), r.get("change_date"),
            r.get("effective_date"), r.get("reason"), r.get("narrative_en"),
            r.get("status"), r.get("announcement_url"), r.get("scraped_at"),
        ])
        _apply_status_fill(ws, i, r.get("status"))

    _autofit(ws)


# ── JIRA Upload Sheet (stub — last priority) ──────────────────────────────────

def _write_jira_upload(wb, fund_commitments, people_moves, watchlist):
    ws = wb.create_sheet("JIRA_Upload")
    # 20-column JIRA bulk import format — stub for now
    cols = [
        "Summary", "Issue Type", "Priority", "Reporter", "Assignee",
        "Labels", "Component", "Description", "Due Date", "Sprint",
        "Epic Link", "Story Points", "Custom 1", "Custom 2", "Custom 3",
        "Custom 4", "Custom 5", "Custom 6", "Custom 7", "Custom 8",
    ]
    _header_row(ws, 1, cols)
    ws.append(["[JIRA Upload format — to be implemented in next phase]"] + [""] * 19)
    _autofit(ws)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _header_row(ws, row_num: int, headers: list[str]):
    ws.row_dimensions[row_num].height = 20
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_body_style(ws, row_num: int):
    for cell in ws[row_num]:
        cell.font = _BODY_FONT


def _apply_status_fill(ws, row_num: int, status: str):
    if status == "NEW":
        fill = _NEW_FILL
    elif status == "CHANGED":
        fill = _CHANGED_FILL
    else:
        return
    for cell in ws[row_num]:
        cell.fill = fill


def _autofit(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))
