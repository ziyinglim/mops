"""
MOPSOV Scraper — Fund Commitments + People Moves
Direct HTTP requests — no browser needed.
POST to: https://mopsov.twse.com.tw/mops/web/ezsearch_query
Detail pages from: https://emops.twse.com.tw/server-java/t05sr01_1_e
"""

import asyncio
import hashlib
import json
import logging
import re
import sys
import threading
import warnings
from http.server import BaseHTTPRequestHandler, HTTPServer

sys.stdout.reconfigure(encoding="utf-8")
from datetime import datetime, timezone, timedelta
from pathlib import Path

import httpx
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("mopsov")

# ── Config ────────────────────────────────────────────────────────────────────

MOPSOV_SEARCH_URL  = "https://mopsov.twse.com.tw/mops/web/ezsearch_query"
EMOPS_HOST         = "https://emops.twse.com.tw"
TWSE_DOC_URL       = "https://doc.twse.com.tw/server-java/t57sb01"
_TYPEK_OPTIONS     = ["sii", "otc", "rotc", "co"]
_EMOPS_PROFILE_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
    "Referer":      f"https://emops.twse.com.tw/server-java/t58query",
    "Origin":       "https://emops.twse.com.tw",
    "Content-Type": "application/x-www-form-urlencoded",
}

OUTPUT_DIR        = Path("output")
ARCHIVE_DIR       = Path("storage/archive")
STATE_DIR         = Path("storage/state")
PDF_DIR           = Path("storage/pdfs")
CHECK_STATE_PATH  = STATE_DIR / "check_states.json"
API_PORT          = 8502   # port for shared review-state sync; must be open on the host server

# ── Shared review-state API ────────────────────────────────────────────────────
_file_lock = threading.Lock()


def _read_states() -> dict:
    if not CHECK_STATE_PATH.exists():
        return {}
    with _file_lock:
        try:
            return json.loads(CHECK_STATE_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}


def _write_states(patch: dict) -> None:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    with _file_lock:
        existing: dict = {}
        if CHECK_STATE_PATH.exists():
            try:
                existing = json.loads(CHECK_STATE_PATH.read_text(encoding="utf-8"))
            except Exception:
                pass
        existing.update(patch)
        CHECK_STATE_PATH.write_text(
            json.dumps(existing, indent=2, ensure_ascii=False), encoding="utf-8"
        )


class _CheckHandler(BaseHTTPRequestHandler):
    def _cors(self) -> None:
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def do_OPTIONS(self) -> None:
        self.send_response(204); self._cors(); self.end_headers()

    def do_GET(self) -> None:
        if self.path.rstrip("/") not in ("/api/checks", ""):
            self.send_response(404); self.end_headers(); return
        body = json.dumps(_read_states(), ensure_ascii=False).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self._cors(); self.end_headers(); self.wfile.write(body)

    def do_POST(self) -> None:
        length = int(self.headers.get("Content-Length", 0))
        try:
            payload = json.loads(self.rfile.read(length))
            _write_states(payload)
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self._cors(); self.end_headers(); self.wfile.write(b'{"ok":true}')
        except Exception as exc:
            self.send_response(500); self._cors(); self.end_headers()
            self.wfile.write(json.dumps({"error": str(exc)}).encode())

    def log_message(self, *args) -> None:
        pass


_api_started = False
_api_lock    = threading.Lock()


def start_api_server(port: int = API_PORT) -> None:
    """Start the check-state sync server in a background daemon thread.

    Safe to call multiple times — only one server is ever started per process.
    Call this from your Streamlit app.py (or any wrapper) before serving the report:

        import mopsov
        mopsov.start_api_server()
    """
    global _api_started
    with _api_lock:
        if _api_started:
            return
        try:
            srv = HTTPServer(("0.0.0.0", port), _CheckHandler)
            threading.Thread(target=srv.serve_forever, daemon=True).start()
            _api_started = True
        except OSError:
            _api_started = True  # already bound (hot-reload)

# Default date range — 2 years back to today
SDATE = (datetime.now() - timedelta(days=730)).strftime("%Y/%m/%d")
EDATE = datetime.now().strftime("%Y/%m/%d")

_SEASON_LABEL   = {1: "Q1", 2: "H1", 3: "Q3", 4: "Q4"}
_FS_SEASON_DATE = {1: "03/31", 2: "06/30", 3: "09/30", 4: "12/31"}

WATCHLIST = [
    # ── firm_id: replace with your internal numeric ID (4-6 digits) ────────────
    # ── firm_url: replace with your internal profile/CRM link ──────────────────
    {"stock_code": "2881", "name_en": "Fubon Financial Holding",                "company_type": "financial holding company",
     "firm_id": "10001", "firm_url": "https://example.com",
     "f26_name_en": "Fubon Life Insurance", "f26_display_code": "28810012",
     "f26_subsidiaries": [
         {"display_code": "28810012", "name_en": "Fubon Life Insurance"},
         {"display_code": "28810006", "name_en": "Taipei Fubon Commercial Bank"},
     ]},
    {"stock_code": "2882", "name_en": "Cathay Financial Holdings",               "company_type": "financial holding company",
     "firm_id": "10002", "firm_url": "https://example.com",
     "f26_subsidiaries": [
         {"display_code": "28820001", "name_en": "Cathay Life Insurance"},
         {"display_code": "28820004", "name_en": "Cathay United Bank"},
     ]},
    {"stock_code": "2891", "name_en": "CTBC Financial Holdings",                 "company_type": "financial holding company",
     "firm_id": "10003", "firm_url": "https://example.com",
     "f26_name_en": "CTBC Bank",
     "f26_subsidiaries": [
         {"display_code": "28910001", "name_en": "CTBC Bank"},
     ]},
    {"stock_code": "2330", "name_en": "Taiwan Semiconductor Manufacturing Company", "company_type": "semiconductor company",
     "firm_id": "10004", "firm_url": "https://example.com"},
    {"stock_code": "5874", "name_en": "Nan Shan Life Insurance",                 "company_type": "insurance company",
     "firm_id": "10005", "firm_url": "https://example.com"},
    {"stock_code": "2317", "name_en": "Foxconn Technology Group",                "company_type": "technology company",
     "firm_id": "10006", "firm_url": "https://example.com"},
    {"stock_code": "2886", "name_en": "Mega Financial Holdings",                 "company_type": "financial holding company",
     "firm_id": "10007", "firm_url": "https://example.com",
     "f26_name_en": "Mega International Commercial Bank",
     "f26_subsidiaries": [
         {"display_code": "28860005", "name_en": "Mega International Commercial Bank"},
     ]},
    {"stock_code": "2880", "name_en": "Hua Nan Commercial Bank",                 "company_type": "commercial bank",
     "firm_id": "10008", "firm_url": "https://example.com"},
    {"stock_code": "5857", "name_en": "Land Bank of Taiwan",                     "company_type": "state-owned bank",
     "firm_id": "10009", "firm_url": "https://example.com"},
    {"stock_code": "2888", "name_en": "Shin Kong Life Insurance",                "company_type": "insurance company",
     "firm_id": "10010", "firm_url": "https://example.com", "delisted": True},
    {"stock_code": "2801", "name_en": "Chang Hwa Bank",                          "company_type": "commercial bank",
     "firm_id": "10011", "firm_url": "https://example.com"},
    {"stock_code": "2890", "name_en": "SinoPac Financial Holdings",               "company_type": "financial holding company",
     "firm_id": "10012", "firm_url": "https://example.com",
     "f26_name_en": "Bank SinoPac",
     "f26_subsidiaries": [
         {"display_code": "28900001", "name_en": "Bank SinoPac"},
     ]},
    {"stock_code": "5876", "name_en": "Shanghai Commercial & Savings Bank",      "company_type": "commercial bank",
     "firm_id": "10013", "firm_url": "https://example.com"},
    {"stock_code": "2885", "name_en": "Yuanta Financial Holdings",               "company_type": "financial holding company",
     "firm_id": "10014", "firm_url": "https://example.com",
     "f26_name_en": "Yuanta Commercial Bank",
     "f26_subsidiaries": [
         {"display_code": "28850007", "name_en": "Yuanta Commercial Bank"},
     ]},
    {"stock_code": "2833", "name_en": "Taiwan Life Insurance",                   "company_type": "insurance company",
     "firm_id": "10015", "firm_url": "https://example.com"},
    {"stock_code": "2867", "name_en": "Mercuries Life Insurance",                "company_type": "insurance company",
     "firm_id": "10016", "firm_url": "https://example.com"},
    {"stock_code": "5873", "name_en": "TransGlobe Life Insurance",               "company_type": "insurance company",
     "firm_id": "10017", "firm_url": "https://example.com"},
    {"stock_code": "2382", "name_en": "Quanta Computer",                         "company_type": "technology company",
     "firm_id": "10018", "firm_url": "https://example.com"},
    {"stock_code": "3231", "name_en": "Wistron Corporation",                     "company_type": "technology company",
     "firm_id": "10019", "firm_url": "https://example.com"},
    {"stock_code": "3711", "name_en": "ASE Technology Holding",                  "company_type": "semiconductor company",
     "firm_id": "10020", "firm_url": "https://example.com"},
    {"stock_code": "5859", "name_en": "Farglory Life Insurance",                 "company_type": "insurance company",
     "firm_id": "10021", "firm_url": "https://example.com"},
    {"stock_code": "2454", "name_en": "MediaTek",                                "company_type": "semiconductor company",
     "firm_id": "10022", "firm_url": "https://example.com"},
    {"stock_code": "2897", "name_en": "O-Bank",                                  "company_type": "digital bank",
     "firm_id": "10023", "firm_url": "https://example.com"},
    {"stock_code": "2002", "name_en": "China Steel Corporation",                 "company_type": "steel company",
     "firm_id": "10024", "firm_url": "https://example.com"},
]
_WATCHLIST_MAP = {w["stock_code"]: w for w in WATCHLIST}

# Subsidiary detection: maps (parent_code, name_pattern) → subsidiary public code.
# Used to re-label FC/PM records where the announcement title names the subsidiary.
_SUBSIDIARY_PATTERNS: dict[str, list[tuple]] = {
    "2881": [
        (re.compile(r"Fubon Life(?:\s+Insurance)?", re.I),          "28810012"),
        (re.compile(r"Taipei Fubon(?:\s+(?:Commercial\s+)?Bank)?", re.I), "28810006"),
    ],
    "2882": [
        (re.compile(r"Cathay United Bank", re.I),                   "28820004"),
        (re.compile(r"Cathay Life(?:\s+Insurance)?", re.I),         "28820001"),
    ],
    "2885": [
        (re.compile(r"Yuanta Commercial Bank", re.I),               "28850007"),
    ],
    "2886": [
        (re.compile(r"Mega International(?:\s+Commercial\s+Bank)?", re.I), "28860005"),
    ],
    "2890": [
        (re.compile(r"Bank SinoPac|SinoPac Bank", re.I),                   "28900001"),
    ],
    "2891": [
        (re.compile(r"CTBC Bank", re.I),                                   "28910001"),
    ],
}

# Maps internal TWSE doc-server subsidiary_code → public-facing subsidiary code (for FS tab).
# Only needed when the holding company files under its own stock_code but the FS PDF is for a
# named subsidiary with a different public display code (e.g. Fubon Financial 2881 → Fubon Life).
# Banks listed directly (2880/2886/2890) have sub_code != stock_code but we want to display
# them under their watchlist code, so they are intentionally omitted — the fallback handles it.
_FS_SUBSIDIARY_CODE_MAP: dict[str, str] = {
    "5865":   "28810012",   # Fubon Life Insurance               (parent: Fubon Financial 2881)
    "5836":   "28810006",   # Taipei Fubon Commercial Bank        (parent: Fubon Financial 2881)
    "5846":   "28820001",   # Cathay Life Insurance              (parent: Cathay Financial 2882)
    "5835":   "28820004",   # Cathay United Bank                 (parent: Cathay Financial 2882)
    "5852":   "28850007",   # Yuanta Commercial Bank             (parent: Yuanta Financial 2885)
    "000700": "28860005",   # Mega International Commercial Bank  (parent: Mega Financial 2886)
    "5843":   "28860005",   # Mega International Commercial Bank  (parent: Mega Financial 2886)
    "5849":   "28900001",   # Bank SinoPac                       (parent: SinoPac Financial 2890)
    "5841":   "28910001",   # CTBC Bank                          (parent: CTBC Financial 2891)
}

def _resolve_subsidiary(parent_code: str, subject: str) -> str:
    """Return the subsidiary public code if the subject names a known subsidiary, else parent_code."""
    for pattern, sub_code in _SUBSIDIARY_PATTERNS.get(parent_code, []):
        if pattern.search(subject):
            return sub_code
    return parent_code


def _fc_resolve_code(stock_code: str, subject: str) -> str:
    """Like _resolve_subsidiary but for fund commitments.

    If the subject can't identify a specific subsidiary and the parent is a
    holding company with a single default investing entity, fall back to that
    entity so FC headlines name the correct fund-investing subsidiary rather
    than the FHC itself.
    """
    resolved = _resolve_subsidiary(stock_code, subject)
    if resolved != stock_code:
        return resolved
    entry = _WATCHLIST_MAP.get(stock_code, {})
    if not _HOLDING_RE.search(entry.get("name_en", "")):
        return resolved
    # Prefer f26_display_code (explicitly designated primary investing entity)
    if entry.get("f26_display_code"):
        return entry["f26_display_code"]
    # Fall back to sole subsidiary display_code
    subs = entry.get("f26_subsidiaries", [])
    if len(subs) == 1:
        return subs[0]["display_code"]
    return resolved  # Multiple subsidiaries — can't pick a default


# Reverse lookup: subsidiary display code → parent WATCHLIST stock_code
_SUBSIDIARY_TO_PARENT: dict[str, str] = {
    sub_code: parent
    for parent, patterns in _SUBSIDIARY_PATTERNS.items()
    for _, sub_code in patterns
}

# Subsidiaries with no standalone EMOPS/MOPSOV profile — shown in Company Profiles with a notice
_SUBSIDIARY_STUBS = [
    {"stock_code": "28810012", "company_name_en": "Fubon Life Insurance",         "no_filing_data": True},
    {"stock_code": "28810006", "company_name_en": "Taipei Fubon Commercial Bank", "no_filing_data": True},
    {"stock_code": "28820001", "company_name_en": "Cathay Life Insurance",        "no_filing_data": True},
    {"stock_code": "28820004", "company_name_en": "Cathay United Bank",           "no_filing_data": True},
    {"stock_code": "28850007", "company_name_en": "Yuanta Commercial Bank",           "no_filing_data": True},
    {"stock_code": "28860005", "company_name_en": "Mega International Commercial Bank", "no_filing_data": True},
    {"stock_code": "28900001", "company_name_en": "Bank SinoPac",                       "no_filing_data": True},
    {"stock_code": "28910001", "company_name_en": "CTBC Bank",                         "no_filing_data": True},
]

# Allowlist: fund commitment must match at least one of these in name+type+statement
_FUND_ALLOWLIST_RE = re.compile(
    r"\bfund\b|fund commitment|private equity|\bP\.?E\b|venture capital|"
    r"real estate|\bREIT\b|infrastructure|hedge fund|"
    r"alternative (?:asset|investment)|secondar(?:y|ies)|"
    r"private (?:credit|debt|market)|mezzanine|"
    r"growth (?:equity|capital)|buyout|\bLBO\b|"
    r"co.?invest|special situation|distressed|"
    r"natural resource|commodit(?:y|ies)|gold|precious metal|"
    r"private fund|real asset|"
    r"direct investment|InvIT|"
    r"\bdeal\b|\bacquisition\b",
    re.IGNORECASE
)

# Denylist: exclude even if allowlist matched
_FUND_EXCLUDE_RE = re.compile(
    r"asset.backed.securiti[sz]ation|mortgage.backed|"
    r"\bdisposal\b|"
    r"common (?:share|stock)|ordinary share|preferred (?:share|stock)|"
    r"\bdebenture|\bcorporate bonds?|\bbond issue|\bnote issue|"
    r"\bfinancial bonds?|\bsecured bonds?|\bunsecured bonds?\b|"
    r"\bfixed.income\b|\bfinancial products?\b|\bcertificate of deposit\b|"
    r"\bshares\b|"
    r"RMB structured deposit|structured deposit|"
    r"(?<!private )(?<!growth )\bequity\b",
    re.IGNORECASE
)

# Standard fund type categories — everything else is normalised to ""
_FUND_TYPE_MAP = [
    ("Private Equity", re.compile(
        r"private equity|\bP\.?E\.?\b|venture capital|\bV\.?C\.?\b|buyout|\bLBO\b|"
        r"growth (?:equity|capital)|mezzanine|co.?invest|special situation|distressed|"
        r"secondar(?:y|ies)|direct investment",
        re.IGNORECASE)),
    ("Real Estate", re.compile(
        r"real estate|\bREIT\b|real asset|property fund",
        re.IGNORECASE)),
    ("Hedge Funds", re.compile(
        r"hedge fund",
        re.IGNORECASE)),
    ("Infrastructure", re.compile(
        r"infrastructure|InvIT",
        re.IGNORECASE)),
    ("Private Debt", re.compile(
        r"private (?:debt|credit|lending|loan)|direct lending",
        re.IGNORECASE)),
    ("Natural Resources", re.compile(
        r"natural resource|commodit(?:y|ies)|energy|mining|oil\b|gas\b|timber|"
        r"agriculture|precious metal|gold",
        re.IGNORECASE)),
]

def _normalize_fund_type(raw_type: str, fund_name: str = "") -> str:
    """Map raw fund_type to one of 6 standard categories, or '' if unrecognised."""
    text = raw_type + " " + fund_name
    for label, pattern in _FUND_TYPE_MAP:
        if pattern.search(text):
            return label
    return ""

_HOLDING_RE = re.compile(r"financial.hold", re.I)

# Fund name looks like a date or date-range (not a real fund name)
_DATE_NAME_RE = re.compile(r"^\d{4}/\d{2}/\d{2}(\s*~\s*\d{4}/\d{2}/\d{2})?$")

# Roles we actively track — sorted longest-first so the regex prefers specific matches
TRACKED_ROLES = sorted([
    "Chairman",
    "Chief Executive Officer",
    "Chief Finance Officer", "Chief Financial Officer",
    "Chief Investment Officer",
    "Chief Operating Officer", "Chief Operations Officer",
    "Chief Technology Officer",
    "Chief Strategy Officer",
    "General Manager",
    "Global Head of Alternative Assets", "Global Head of Alternative Investments",
    "Global Head of Alternatives", "Global Head of Hedge Funds",
    "Global Head of Infrastructure", "Global Head of Investor Relations",
    "Global Head of Private Debt", "Global Head of Private Equity",
    "Global Head of Real Assets", "Global Head of Real Estate",
    "Global Head of Secondaries",
    "Head of Alternative Assets", "Head of Alternative Investment",
    "Head of Alternative Investments", "Head of Alternatives",
    "Head of Asset Allocation", "Head of Asset Management",
], key=len, reverse=True)

_ROLE_ABBREV = {
    "CEO": "Chief Executive Officer", "CFO": "Chief Financial Officer",
    "CIO": "Chief Investment Officer", "COO": "Chief Operating Officer",
    "CTO": "Chief Technology Officer", "CSO": "Chief Strategy Officer",
}

_ROLE_FULL_TO_ABBREV = {
    "Chief Executive Officer": "CEO",
    "Chief Financial Officer": "CFO",
    "Chief Finance Officer": "CFO",
    "Chief Investment Officer": "CIO",
    "Chief Operating Officer": "COO",
    "Chief Operations Officer": "COO",
    "Chief Technology Officer": "CTO",
    "Chief Strategy Officer": "CSO",
    "Chairman": "Chair",
    "General Manager": "GM",
    "Head of Alternative Investments": "Head of Alts",
    "Head of Alternative Assets": "Head of Alts",
    "Head of Alternatives": "Head of Alts",
    "Global Head of Alternative Investments": "Global Head of Alts",
    "Global Head of Alternative Assets": "Global Head of Alts",
    "Global Head of Alternatives": "Global Head of Alts",
}

_TRACKED_ROLES_RE = re.compile(
    "|".join(re.escape(r) for r in TRACKED_ROLES) + r"|\b(CEO|CFO|CIO|COO|CTO|CSO)\b",
    re.IGNORECASE,
)

# Broad keywords used to decide which MOPS announcements to fetch
PEOPLE_KEYWORDS = (
    TRACKED_ROLES
    + list(_ROLE_ABBREV.keys())
    + ["執行長", "投資長", "財務長", "總經理", "董事長", "President"]
)

_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://mopsov.twse.com.tw/mops/web/ezsearch",
    "Origin": "https://mopsov.twse.com.tw",
    "Content-Type": "application/x-www-form-urlencoded",
}

# ── HTTP Search ───────────────────────────────────────────────────────────────

async def search_mopsov(stock_code: str, pro_item: str, sdate: str = None) -> list[dict]:
    """POST to ezsearch_query and return list of announcement rows."""
    data = {
        "step": "00",
        "RADIO_CM": "2",
        "TYPEK": "CO_MARKET",
        "CO_ID": stock_code,
        "PRO_ITEM": pro_item,
        "SUBJECT": "",
        "SDATE": sdate or SDATE,
        "EDATE": EDATE,
        "lang": "EN",
    }
    async with httpx.AsyncClient(headers=_HEADERS, timeout=30, verify=False) as client:
        try:
            resp = await client.post(MOPSOV_SEARCH_URL, data=data)
            resp.raise_for_status()
            logger.info("Search [%s/%s] → %d chars", stock_code, pro_item, len(resp.text))
            return _parse_results(resp.text, stock_code)
        except Exception as exc:
            logger.error("Search failed [%s/%s]: %s", stock_code, pro_item, exc)
            return []


def _parse_results(response_text: str, stock_code: str) -> list[dict]:
    try:
        # Strip UTF-8 BOM and leading whitespace before the JSON object
        clean = response_text.lstrip("\ufeff \r\n\t")
        data = json.loads(clean)
        records = data.get("data", [])
    except Exception:
        logger.error("Failed to parse JSON response for %s", stock_code)
        return []

    rows = []
    for item in records:
        subject = item.get("SUBJECT", "").replace("\r\n", " ").replace("\n", " ").strip()
        rows.append({
            "date": item.get("CDATE", ""),
            "stock_code": item.get("COMPANY_ID", stock_code),
            "subject": subject,
            "url": item.get("HYPERLINK", ""),
        })
    logger.info("Parsed %d records for %s", len(rows), stock_code)
    return rows


# ── Detail Page Parser ────────────────────────────────────────────────────────

async def fetch_detail(url: str, retries: int = 3) -> str:
    """Fetch announcement detail page HTML with retries and rate-limit handling."""
    if not url:
        return ""
    for attempt in range(1, retries + 1):
        try:
            async with httpx.AsyncClient(timeout=30, verify=False) as client:
                resp = await client.get(url)
                resp.raise_for_status()
                html = ""
                for enc in ("utf-8", "cp950", "big5"):
                    try:
                        html = resp.content.decode(enc)
                        break
                    except UnicodeDecodeError:
                        continue
                if not html:
                    html = resp.content.decode("utf-8", errors="replace")
                # Detect server rate-limit page (查詢過量 = "Too many queries")
                if "查詢過量" in html:
                    wait = 10 * attempt
                    logger.warning("Rate limited by server — waiting %ds before retry %d/%d [%s]", wait, attempt, retries, url)
                    await asyncio.sleep(wait)
                    continue
                return html
        except Exception as exc:
            if attempt < retries:
                logger.warning("Detail fetch attempt %d/%d failed [%s]: %s — retrying", attempt, retries, url, exc)
                await asyncio.sleep(2 ** attempt)
            else:
                logger.error("Detail fetch failed after %d attempts [%s]: %s", retries, url, exc)
    return ""


def extract_statement(html: str) -> str:
    """Extract the Statement column content from detail page HTML."""
    soup = BeautifulSoup(html, "lxml")
    header = soup.find("td", string=re.compile(r"Statement", re.I))
    if header:
        sibling = header.find_next_sibling("td")
        if sibling:
            return sibling.get_text("\n", strip=True)
    # Fallback: find the largest td containing numbered fields
    best = ""
    for td in soup.find_all("td"):
        text = td.get_text("\n", strip=True)
        if re.search(r"^\d+\.", text, re.MULTILINE) and len(text) > len(best):
            best = text
    return best


def parse_statement_fields(text: str) -> dict[int, str]:
    """Split numbered field blocks e.g. '1. Fund name...' → {1: 'Fund name...'}"""
    text = re.sub(r"\r\n|\r", "\n", text)
    parts = re.split(r"(?:^|\n)(\d{1,2})\.\s*", text, flags=re.MULTILINE)
    fields: dict[int, str] = {}
    i = 1
    while i < len(parts) - 1:
        try:
            fields[int(parts[i])] = parts[i + 1].strip()
        except (ValueError, IndexError):
            pass
        i += 2
    return fields


# ── Fund Commitments ──────────────────────────────────────────────────────────

async def scrape_fund_commitments(stock_code: str, sdate: str = None) -> list[dict]:
    rows = []
    for pro_item in ["M20", "M24"]:
        rows.extend(await search_mopsov(stock_code, pro_item, sdate=sdate))
    results = []
    for row in rows:
        if not _FUND_ALLOWLIST_RE.search(row["subject"]):
            continue
        await asyncio.sleep(0.8)   # throttle to avoid server rate limit
        html = await fetch_detail(row["url"])
        if not html:
            continue
        statement = extract_statement(html)
        if not statement:
            logger.warning("Empty statement: %s", row["url"])
            continue
        fields = parse_statement_fields(statement)
        f1, f2 = fields.get(1, ""), fields.get(2, "")

        # Strip the standard field label prefix, keep only fund name and type
        f1_clean = f1.rsplit(":", 1)[-1].strip() if ":" in f1 else f1
        parts = [p.strip() for p in f1_clean.split(";")]
        fund_name = _clean_fund_name(parts[0] if parts else f1_clean)
        raw_fund_type = parts[1] if len(parts) > 1 else ""

        # Skip disposal/secondary transactions and unparseable names
        if re.search(r"^The partnership interests? of\b", fund_name, re.I):
            continue
        if _DATE_NAME_RE.match(fund_name):   # field 1 parsed as a date, not a fund name
            continue

        date_match = re.search(r"\d{4}/\d{2}/\d{2}", f2)

        # Find the amount field — field number varies across M20/M24 formats
        amount_field = next(
            (v for v in fields.values()
             if re.search(r"total monetary amount of the transaction", v, re.I)),
            None
        )
        # Skip if no standard fund commitment amount field found (e.g. property/lease)
        if not amount_field:
            continue

        total_match = re.search(r"\nTotal monetary amount[^:]*:\s*([^\n]+)", amount_field, re.I)
        if total_match:
            amount_raw = total_match.group(1).strip()
        else:
            amount_parts = [p.strip() for p in amount_field.split(";")]
            amount_raw = next((p for p in reversed(amount_parts)
                               if p and not re.match(r"^N/?A\b", p, re.I)), "")
        # Strip update/revision annotations from raw amount text
        amount_raw = re.sub(r"\s*\(updated\)", "", amount_raw, flags=re.IGNORECASE).strip()
        # Deduplicate leading currency ticker (e.g. "EUR EUR 50,000,000" → "EUR 50,000,000")
        amount_raw = re.sub(r"^([A-Z]{3})\s+\1\b", r"\1", amount_raw)
        currency_match = re.search(r"(?<![A-Za-z])(USD|EUR|GBP|JPY|CNY|RMB|TWD|HKD|SGD|AUD|CAD)(?![A-Za-z])", amount_raw)

        searchable = fund_name + " " + raw_fund_type + " " + row["subject"] + " " + statement
        if not _FUND_ALLOWLIST_RE.search(searchable):
            continue
        if _FUND_EXCLUDE_RE.search(fund_name + " " + raw_fund_type + " " + row["subject"]):
            continue

        resolved_code = _fc_resolve_code(stock_code, row["subject"])
        _, bs_date = _get_latest_aum(resolved_code)
        raw_currency = currency_match.group(1) if currency_match else ""
        if raw_currency == "RMB":
            raw_currency = "CNY"
        formatted_amount, fx_url = await _format_commitment_amount(amount_raw, raw_currency)
        fund_type = _normalize_fund_type(raw_fund_type, fund_name)
        headline = _build_fund_headline(resolved_code, fund_name, formatted_amount, fund_type) if fund_type else ""
        twd_match = re.search(r"TWD ([\d,]+) million", formatted_amount)
        twd_amount_mn = twd_match.group(1) if twd_match else ""
        commit_date = date_match.group(0) if date_match else ""
        key_events = _build_fc_key_event(
            resolved_code,
            commit_date or row["date"], fund_name, formatted_amount, fund_type
        ) if fund_type else ""

        results.append({
            "stock_code": resolved_code,
            "announcement_date": row["date"],
            "subject": row["subject"],
            "headline": headline,
            "key_events": key_events,
            "fund_name": fund_name,
            "fund_type": fund_type,
            "commitment_date": commit_date,
            "commitment_amount_raw": amount_raw,
            "twd_amount_mn": twd_amount_mn,
            "commitment_currency": raw_currency,
            "commitment_amount_numeric": _parse_amount(amount_raw),
            "fx_url": fx_url,
            "bs_date": bs_date,
            "url": row["url"],
        })
    logger.info("Fund commitments [%s]: %d found", stock_code, len(results))
    return results


# ── People Moves ──────────────────────────────────────────────────────────────

async def scrape_people_moves(stock_code: str, sdate: str = None) -> list[dict]:
    results = []
    seen_urls: set[str] = set()
    for pro_item in ["B02", "M08"]:
        rows = await search_mopsov(stock_code, pro_item, sdate=sdate)
        for row in rows:
            if row.get("url") in seen_urls:
                continue
            seen_urls.add(row.get("url", ""))
            if not _matches(row["subject"], PEOPLE_KEYWORDS):
                continue
            await asyncio.sleep(0.8)
            html = await fetch_detail(row["url"])
            if not html:
                continue
            statement = extract_statement(html)
            if not statement:
                continue
            fields = parse_statement_fields(statement)

            def clean(text):
                if text and ":" in text and len(text.split(":", 1)[0]) > 10:
                    return text.split(":", 1)[1].strip()
                return (text or "").strip()

            role_type      = clean(fields.get(1, ""))
            change_date    = _extract_date(fields.get(2, ""))
            prev_holder    = clean(fields.get(3, ""))
            new_holder     = clean(fields.get(4, ""))
            change_type    = clean(fields.get(5, "")).lower()
            reason         = clean(fields.get(6, ""))
            effective_date = _extract_date(fields.get(7, ""))

            # Primary: extract role from field 1 suffix (text after last ':')
            # e.g. "...representatives):COO" → "COO" → "Chief Operating Officer"
            role_raw = role_type.rsplit(":", 1)[-1].strip() if ":" in role_type else ""
            role_title = (_extract_tracked_role(role_raw)
                          or _extract_tracked_role(new_holder)
                          or _extract_tracked_role(prev_holder))
            if not role_title:
                continue

            new_holder_clean  = _clean_holder_name(new_holder)
            prev_holder_clean = _clean_holder_name(prev_holder)

            resolved_code = _resolve_subsidiary(stock_code, row["subject"])
            narrative = _NARRATIVE_ABBREV_RE.sub(
                "", _build_narrative(resolved_code, role_title,
                                     new_holder_clean, prev_holder_clean,
                                     change_type, effective_date)
            ).strip()
            _, bs_date = _get_latest_aum(resolved_code)
            key_events = _build_pm_key_event(
                resolved_code, role_title, new_holder_clean, prev_holder_clean,
                effective_date or change_date or row["date"], change_type
            )
            results.append({
                "stock_code": resolved_code,
                "announcement_date": row["date"],
                "subject": row["subject"],
                "role_type": role_type,
                "role_title": role_title,
                "change_date": change_date,
                "previous_holder": prev_holder_clean,
                "new_holder": new_holder_clean,
                "change_type": change_type,
                "reason": reason,
                "effective_date": effective_date,
                "narrative_en": narrative,
                "key_events": key_events,
                "bs_date": bs_date,
                "url": row["url"],
            })
    logger.info("People moves [%s]: %d found", stock_code, len(results))
    return results


# ── Helpers ───────────────────────────────────────────────────────────────────

_NARRATIVE_ABBREV_RE = re.compile(
    r",?\s*\bCo\.,?\s*Ltd\.?(?!\w)|,?\s*\bLtd\.?(?!\w)|,?\s*\bInc\.?(?!\w)|,?\s*\bCorp\.?(?!\w)",
    re.IGNORECASE,
)

_LEGAL_SUFFIX_RE = re.compile(
    r"[,\s]*\b("
    r"S\.C\.S\.,?\s*SICAV-RAIF|SICAV-RAIF|S\.C\.S\.|SCSp|S\.C\.A\.|RAIF|SICAV|SCA|"
    r"L\.P\.|L\.P|(?<!\w)LP(?!\w)|"
    r"Ltd\.|Ltd|Limited|"
    r"LLC|L\.L\.C\.|"
    r"Pte\.?\s*Ltd\.?|(?<!\w)Pte(?!\w)|"
    r"Inc\.|(?<!\w)Inc(?!\w)|"
    r"Corp\.|(?<!\w)Corp(?!\w)|"
    r"(?<!\w)Co\.(?!\w)|"
    r"GmbH|S\.A\.|(?<!\w)SA(?!\w)|B\.V\.|(?<!\w)BV(?!\w)|N\.V\.|(?<!\w)NV(?!\w)|"
    r"Sàrl|SARL"
    r")\s*$",
    re.IGNORECASE,
)

def _clean_fund_name(name: str) -> str:
    """Strip trailing legal entity suffixes (L.P., Ltd., SCSp, SICAV-RAIF, etc.) and punctuation."""
    prev = None
    while prev != name:
        prev = name
        name = _LEGAL_SUFFIX_RE.sub("", name).strip().rstrip(".,").strip()
    return name

def _extract_tracked_role(text: str) -> str:
    """Return the best-matching tracked role title from holder text (searches after '/')."""
    if not text:
        return ""
    # Prefer the section after '/' where the role description lives
    parts = re.split(r"\s*/\s*", text, maxsplit=1)
    search_in = parts[1] if len(parts) > 1 else text
    best_raw_len = 0
    best = ""
    for m in _TRACKED_ROLES_RE.finditer(search_in):
        matched = m.group(0)
        # Skip abbreviations that appear inside "to/of the CEO" constructions
        if len(matched) <= 3 and re.search(
            r'\b(?:to|of)\s+the\s+' + re.escape(matched), search_in, re.I
        ):
            continue
        expanded = _ROLE_ABBREV.get(matched.upper(), matched)
        # Score by raw match length — full titles beat abbreviations
        if len(matched) > best_raw_len:
            best_raw_len = len(matched)
            best = expanded
    return best.strip()

# Titles/roles that follow a person's name — strip from first match onwards.
# Ordered longest-first so alternation picks the most specific phrase.
_TITLE_STRIP_RE = re.compile(
    r"\b(?:Senior\s+Executive\s+Vice\s+President"
    r"|Executive\s+Vice\s+President"
    r"|Senior\s+Vice\s+President"
    r"|Vice[\s-]+President"
    r"|Senior\s+Managing\s+Director"
    r"|Managing\s+Director"
    r"|Senior\s+Executive\s+Director"
    r"|Executive\s+Director"
    r"|Independent\s+Director"
    r"|Non-?Executive\s+Director"
    r"|Senior\s+Director"
    r"|Deputy\s+General\s+Manager"
    r"|Assistant\s+General\s+Manager"
    r"|General\s+Manager"
    r"|Chief\s+\w+\s+Officer"
    r"|Senior\s+Executive"
    r"|Executive\s+Officer"
    r"|President"
    r"|Chairman"
    r"|Director"
    r"|Manager"
    r"|C(?:EO|FO|OO|TO|SO|IO|LO|CRO)\b"
    r")\b.*$",
    re.IGNORECASE | re.DOTALL,
)

def _clean_holder_name(text: str) -> str:
    """Extract only the person's name from holder text.

    Handles three formats:
      'SURNAME,GIVEN-NAME'              → joined to 'Surname Given-Name' (Chinese format)
      'Name, Title, Company...'         → takes text before first comma (English comma-separated)
      'Name Senior Vice President Co.'  → strips from first recognised title keyword onward
    """
    if not text or text.lower().strip() in ("none", "nil", "n/a", "na", ""):
        return ""
    # Strip slash-separated role suffix (e.g. "Name / CEO")
    name = re.split(r"\s*/\s*", text.strip(), maxsplit=1)[0].strip()

    if ", " in name or ",\t" in name:
        # English comma-separated: "Suwat Chritamara, CSO, Land and Houses Bank..."
        name = name.split(",")[0].strip()
    elif "," in name:
        # Chinese SURNAME,GIVEN-NAME format — join with space
        name = name.replace(",", " ").strip()

    # Always strip recognised title words (catches both no-comma and post-comma-split cases)
    # e.g. "Ching-Li Chang Senior Executive Vice President Cathay United Bank" → "Ching-Li Chang"
    # e.g. "Sajiv Dalal President Of Tsmc North America" → "Sajiv Dalal"
    name = _TITLE_STRIP_RE.sub("", name).strip()

    name = name.rstrip(".,、，").strip()
    if not name or name.lower() in ("none", "nil", "n/a", "na"):
        return ""
    return name.title()

def _matches(text: str, keywords: list[str]) -> bool:
    t = text.lower()
    return any(k.lower() in t for k in keywords)

def _extract_date(text: str) -> str:
    m = re.search(r"\d{4}/\d{2}/\d{2}|\d{4}-\d{2}-\d{2}", text or "")
    return m.group(0) if m else ""

def _parse_amount(text: str) -> float | None:
    m = re.search(r"[\d,]+(?:\.\d+)?", text.replace(" ", ""))
    if m:
        try:
            return float(m.group().replace(",", ""))
        except ValueError:
            pass
    return None

_FX_CACHE: dict[str, dict] = {}

async def _get_fx_rates(base: str) -> dict[str, float]:
    if base in _FX_CACHE:
        return _FX_CACHE[base]
    try:
        async with httpx.AsyncClient(timeout=10, verify=False) as client:
            resp = await client.get(f"https://open.er-api.com/v6/latest/{base}")
            data = resp.json()
            if data.get("result") == "success":
                _FX_CACHE[base] = data.get("rates", {})
                return _FX_CACHE[base]
    except Exception as exc:
        logger.warning("FX fetch failed [%s]: %s", base, exc)
    return {}

async def _format_commitment_amount(amount_raw: str, orig_currency: str) -> tuple[str, str]:
    """Returns (formatted_amount_str, fx_url).
    formatted_amount like 'TWD 1,846 million (EUR 50 million)'.
    fx_url is an XE.com link for the conversion, empty if no conversion needed."""
    if not amount_raw or re.match(r"^N/?A\b", amount_raw, re.I):
        return "", ""
    # Infer currency from symbol if not explicitly provided (e.g. "US$" → USD, "€" → EUR)
    if not orig_currency:
        sym_map = {"US$": "USD", "NT$": "TWD", "€": "EUR", "£": "GBP", "¥": "JPY", "HK$": "HKD", "A$": "AUD", "S$": "SGD", "C$": "CAD"}
        for sym, cur in sym_map.items():
            if sym in amount_raw:
                orig_currency = cur
                break
    num_match = re.search(r"\d[\d,]*(?:\.\d+)?", amount_raw.replace(" ", ""))
    if not num_match:
        return amount_raw, ""
    amount = float(num_match.group().replace(",", ""))
    # Adjust for amounts already expressed in millions/billions
    if re.search(r"\bbillion\b", amount_raw, re.I):
        amount *= 1e9
    elif re.search(r"\bmillion\b", amount_raw, re.I):
        amount *= 1e6
    orig_millions = amount / 1e6

    if not orig_currency or orig_currency == "TWD":
        return f"TWD {orig_millions:,.0f} million", ""

    orig_str = f"{orig_currency} {orig_millions:,.0f} million"
    fx_url = f"https://www.xe.com/currencyconverter/convert/?Amount={amount:,.0f}&From={orig_currency}&To=TWD"
    rates = await _get_fx_rates(orig_currency)
    twd_rate = rates.get("TWD", 0)
    if twd_rate:
        twd_millions = (amount * twd_rate) / 1e6
        return f"TWD {twd_millions:,.0f} million ({orig_str})", fx_url
    return orig_str, fx_url

def _firm_display_name(stock_code: str) -> str:
    """Return the public display name for a stock code, handling subsidiaries."""
    entry = _WATCHLIST_MAP.get(stock_code)
    if entry:
        return entry.get("name_en", stock_code)
    stub = next((s for s in _SUBSIDIARY_STUBS if s["stock_code"] == stock_code), None)
    return stub["company_name_en"] if stub else stock_code

def _build_fund_headline(stock_code, fund_name, formatted_amount, fund_type=""):
    firm_name = _firm_display_name(stock_code)
    aum, _ = _get_latest_aum(stock_code)
    ft_lower = fund_type.lower() if fund_type else "alternative"
    title = f"{firm_name} commits to new {ft_lower} fund"
    ref = f"The {aum} [firm type]" if aum else "[firm type]"
    committed = f"has committed {formatted_amount} to" if formatted_amount else "has committed to"
    body = f"{ref} {committed} {fund_name}."
    return f"{title}\n\n{body}"

def _build_narrative(stock_code, role, new_holder, prev_holder, change_type, effective_date):
    firm_name = _firm_display_name(stock_code)
    aum, _ = _get_latest_aum(stock_code)
    company_ref = f"The {aum} [firm type]" if aum else "[firm type]"
    date_str = _format_date(effective_date)
    eff = f", effective {date_str}" if date_str else ""
    role_abbrev = _ROLE_FULL_TO_ABBREV.get(role, role)

    has_new  = bool(new_holder  and new_holder.lower()  not in ("none", "nil", "n/a", "na", ""))
    has_prev = bool(prev_holder and prev_holder.lower() not in ("none", "nil", "n/a", "na", ""))
    ct = change_type or ""

    if has_new:
        title = f"{firm_name} appoints new {role_abbrev}"
        body = f"{company_ref} has appointed {new_holder} as {role}{eff}."
        if has_prev:
            surname = new_holder.split()[-1]
            body += f" {surname} will succeed {prev_holder}."
    elif "position adjustment" in ct:
        title = f"{firm_name} {role_abbrev} position eliminated"
        body = f"{company_ref}'s {role} position has been eliminated{eff}."
        if has_prev:
            body += f" {prev_holder} vacated the role."
    elif has_prev:
        if "retirement" in ct:
            title = f"{firm_name}'s {role_abbrev} retires"
            body = f"{prev_holder} stepped down from {role} due to retirement{eff}."
        else:
            title = f"{firm_name}'s {role_abbrev} steps down"
            body = f"{company_ref}'s {role}, {prev_holder}, has stepped down{eff}."
    else:
        title = f"{firm_name} announces {role_abbrev} change"
        body = f"{company_ref} has announced a change in its {role}{eff}."

    return f"{title}\n\n{body}"

_AUM_CACHE: dict[str, tuple[str, str]] = {}

def _populate_aum_cache(balance_history: list[dict]) -> None:
    """Build _AUM_CACHE from in-memory balance history (avoids redundant file I/O).

    For financial holding companies the AUM must reflect the group's *consolidated*
    total assets, not a single subsidiary's figure.  Consolidated records are
    identified by subsidiary_code == parent stock_code (the holding company filed
    the report under its own TWSE doc code, usually for Q1 and Q3).  These are
    prioritised; subsidiary individual records are only used as a fallback.
    """
    by_key: dict[str, list[dict]] = {}
    for rec in balance_history:
        sub_code = rec.get("subsidiary_code", "")
        display_code = _FS_SUBSIDIARY_CODE_MAP.get(sub_code) or rec.get("stock_code", "")
        parent_code = rec.get("stock_code", "")
        for key in {display_code, parent_code}:
            if key:
                by_key.setdefault(key, []).append(rec)

    for key, records in by_key.items():
        entry = _WATCHLIST_MAP.get(key, {})
        is_fhc = bool(_HOLDING_RE.search(entry.get("name_en", "")))

        if is_fhc:
            # Prefer consolidated records (those where the holding company filed
            # under its own stock code: subsidiary_code == parent stock_code).
            # These records may pre-date the roc_year/season fields, so sort by
            # period string which is lexicographically correct (e.g. "2025/Q3").
            own = [r for r in records if str(r.get("subsidiary_code", "")) == key]
            if own:
                own.sort(key=lambda r: r.get("period", ""), reverse=True)
                use = own
            else:
                # No consolidated record available — fall back to subsidiary data
                records.sort(key=lambda r: (r.get("roc_year", 0), r.get("season", 0)), reverse=True)
                use = records
        else:
            records.sort(key=lambda r: (r.get("roc_year", 0), r.get("season", 0)), reverse=True)
            use = records

        for rec in use:
            total_k = rec.get("total_assets_numeric")
            if total_k is None:
                continue
            total = total_k * 1000
            if total >= 1e9:
                _AUM_CACHE[key] = (f"TWD {total/1e9:,.0f} billion", rec.get("period", ""))
            elif total >= 1e6:
                _AUM_CACHE[key] = (f"TWD {total/1e6:,.0f} million", rec.get("period", ""))
            break

def _get_latest_aum(stock_code: str) -> tuple[str, str]:
    """Returns (aum_string, balance_sheet_period) for the given stock code.
    Each entity is looked up by its own code only — no parent fallback.
    If the subsidiary's AUM is missing or unreadable, returns ('', '') so
    headlines omit the figure rather than inheriting a wrong parent value."""
    if _AUM_CACHE:
        return _AUM_CACHE.get(stock_code, ("", ""))
    # File-based fallback used before cache is populated (standalone run).
    # Applies the same FHC-preference logic as _populate_aum_cache so that
    # parent holding-company codes return the group consolidated figure, not
    # whichever subsidiary happens to have the highest roc_year/season.
    path = STATE_DIR / f"{stock_code}_balance_history.json"
    if not path.exists():
        return "", ""
    try:
        records = json.loads(path.read_text(encoding="utf-8"))
        entry  = _WATCHLIST_MAP.get(stock_code, {})
        is_fhc = bool(_HOLDING_RE.search(entry.get("name_en", "")))
        if is_fhc:
            own = [r for r in records if str(r.get("subsidiary_code", "")) == stock_code]
            if own:
                own.sort(key=lambda r: r.get("period", ""), reverse=True)
                records = own
            else:
                records.sort(key=lambda r: (r.get("roc_year", 0), r.get("season", 0)), reverse=True)
        else:
            records.sort(key=lambda r: (r.get("roc_year", 0), r.get("season", 0)), reverse=True)
        for rec in records:
            total_k = rec.get("total_assets_numeric")
            if total_k is None:
                continue
            period = rec.get("period", "")
            total  = total_k * 1000
            if total >= 1e9:
                return f"TWD {total/1e9:,.0f} billion", period
            if total >= 1e6:
                return f"TWD {total/1e6:,.0f} million", period
    except Exception:
        pass
    return "", ""

# ── TWSE Document Server — Quarterly Financial Report PDFs ───────────────────

_TWSE_DOC_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer":    "https://doc.twse.com.tw/",
    "Accept":     "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

def _twse_doc_url(co_id: str, roc_year: int, seamon: int | str = "") -> str:
    return f"{TWSE_DOC_URL}?step=1&colorchg=1&co_id={co_id}&year={roc_year}&seamon={seamon}&mtype=A&"

_SET_COID_RE = re.compile(r"setCoid\(['\"](\w+)['\"]\)", re.I)
_CO_ID_HREF_RE = re.compile(r"co_id[='\"\s]*(['\"]?)(\w+)\1")

def _find_twse_subsidiary(
    soup: BeautifulSoup, target_display_code: str | None = None
) -> tuple[str | None, str | None]:
    """From t57sb01 parent listing page, return (internal_co_id, display_code) for relevant subsidiary.
    Handles both javascript:setCoid('5846') and href co_id= link patterns.
    If target_display_code provided (e.g. '28810012'), matches by display code; else returns first candidate."""
    candidates = []
    for row in soup.find_all("tr"):
        cells = row.find_all(["td", "th"])
        if len(cells) < 2:
            continue
        co_id_str = None
        for cell in cells:
            for tag in cell.find_all(True):
                src = " ".join(str(v) for v in tag.attrs.values() if isinstance(v, str))
                m = _SET_COID_RE.search(src)
                if m:
                    co_id_str = m.group(1)
                    break
                m2 = _CO_ID_HREF_RE.search(src)
                if m2:
                    co_id_str = m2.group(2)
                    break
            if co_id_str:
                break
        if not co_id_str:
            continue
        display_code = cells[0].get_text(strip=True)
        candidates.append((co_id_str, display_code))

    if target_display_code:
        for cid, dc in candidates:
            if dc == target_display_code:
                return cid, dc
    return (candidates[0][0], candidates[0][1]) if candidates else (None, None)


# Matches both AI1 (consolidated) and AI3 (individual) TWSE PDF filenames.
# Format: {gregorian_year}{season}_{co_id}_{AI_type}.pdf  e.g. 202404_2002_AI3.pdf
_AI_FILENAME_RE = re.compile(r"(\d{4})(\d{2})_(\d+)_(AI[13]|AIA)\.pdf", re.I)
_READFILE2_RE   = re.compile(
    r"readfile2\s*\(\s*['\"]([^'\"]+)['\"]\s*,\s*['\"]([^'\"]+)['\"]\s*,\s*['\"]([^'\"]+)['\"]\s*\)",
    re.I
)

def _extract_ai3_links(soup: BeautifulSoup) -> list[tuple[str, str, str, int, int, str]]:
    """Return list of (filename, kind, co_id, roc_year, season, report_type) for AI1/AI3/AIA links.
    report_type is 'AI1' (Chinese consolidated), 'AIA' (English consolidated), or 'AI3' (individual).
    Parses javascript:readfile2('A','5846','202404_5846_AI3.pdf') href patterns used by TWSE doc site."""
    results, seen = [], set()
    for tag in soup.find_all(True):
        src = " ".join(str(v) for v in tag.attrs.values() if isinstance(v, str))
        src += " " + tag.get_text(strip=True)

        m = _READFILE2_RE.search(src)
        if m:
            kind, co_id, filename = m.group(1), m.group(2), m.group(3)
            if not _AI_FILENAME_RE.search(filename):
                continue
        else:
            fm = _AI_FILENAME_RE.search(src)
            if not fm:
                continue
            filename = fm.group(0)
            co_id_m  = re.search(r"_(\d+)_(AI[13]|AIA)", filename, re.I)
            kind, co_id = "A", (co_id_m.group(1) if co_id_m else "")

        if filename in seen:
            continue
        seen.add(filename)
        fm = _AI_FILENAME_RE.search(filename)
        roc_year    = int(fm.group(1)) - 1911
        season      = int(fm.group(2))
        report_type = fm.group(4).upper()
        results.append((filename, kind, co_id, roc_year, season, report_type))
    return results


async def _get_twse_pdf_url(client, kind: str, co_id: str, filename: str) -> str:
    """POST to t57sb01 step=9 with the form params from readfile2() to get the temporary PDF URL."""
    try:
        await asyncio.sleep(1.5)
        r = await client.post(
            "https://doc.twse.com.tw/server-java/t57sb01",
            data={"step": "9", "kind": kind, "co_id": co_id, "filename": filename, "DEBUG": ""},
        )
        html = r.content.decode("big5", errors="replace")
        m = re.search(r"href='(/pdf/[^']+\.pdf)'", html)
        if m:
            return f"https://doc.twse.com.tw{m.group(1)}"
        logger.warning("No PDF URL in step=9 response for %s", filename)
    except Exception as e:
        logger.warning("PDF URL fetch [%s]: %s", filename, e)
    return ""


def _filing_month_to_season(filing_year: int, filing_month: int) -> tuple[int, int]:
    """Map a PDF filing year/month to (roc_year, season).
    Q4 annual: filed Jan–Apr → season=4 of previous calendar year
    Q1:        filed May–Jun → season=1 of same year
    Q2/H1:     filed Jul–Sep → season=2 of same year
    Q3:        filed Oct–Dec → season=3 of same year
    """
    if filing_month <= 4:
        return filing_year - 1912, 4
    elif filing_month <= 6:
        return filing_year - 1911, 1
    elif filing_month <= 9:
        return filing_year - 1911, 2
    else:
        return filing_year - 1911, 3


_ASSET_LABEL_ZH_RE  = re.compile(r"資\s*[産產]\s*[總合]\s*計")
_ASSET_LABEL_EN_RE  = re.compile(r"^\s*Total\s+[Aa]ssets\b")

def _extract_total_assets_from_pdf(pdf_path: Path) -> int | None:
    """Extract total assets from a quarterly financial report PDF.
    Tries both triggers — [資産總計] (Chinese) and [Total Assets] (English) — on every PDF."""
    try:
        import pdfplumber
    except ImportError:
        logger.error("pdfplumber not installed — run: pip install pdfplumber")
        return None

    def _try_chinese(pdf) -> int | None:
        for page in pdf.pages[3:15]:
            for table in (page.extract_tables() or []):
                for row in table:
                    cells = [str(c or "").strip() for c in row]
                    norm_cells = [c.replace(" ", "") for c in cells]
                    if any("資產總計" in c or "資産總計" in c
                           or "資產合計" in c or "資産合計" in c for c in norm_cells):
                        for cell in cells:
                            clean = re.sub(r"[,\s]", "", cell)
                            if re.match(r"^\d{6,}$", clean):
                                return int(clean)
            text = page.extract_text() or ""
            for line in text.split("\n"):
                if not _ASSET_LABEL_ZH_RE.search(line):
                    continue
                if "流動" in line.replace(" ", ""):
                    continue
                norm_line = re.sub(r"(\d) (\d{1,3},)", r"\1\2", line)
                m = re.search(r"\$?\s*([\d,]{6,})", norm_line)
                if m:
                    return int(m.group(1).replace(",", ""))
        return None

    def _try_english(pdf) -> int | None:
        for page in pdf.pages[3:15]:
            for table in (page.extract_tables() or []):
                for row in table:
                    cells = [str(c or "").strip() for c in row]
                    if any(_ASSET_LABEL_EN_RE.match(c) for c in cells):
                        if not any(re.search(r"current|non.current", c, re.I) for c in cells):
                            for cell in cells:
                                clean = re.sub(r"[,\s]", "", cell)
                                if re.match(r"^\d{6,}$", clean):
                                    return int(clean)
            text = page.extract_text() or ""
            for line in text.split("\n"):
                if not _ASSET_LABEL_EN_RE.match(line):
                    continue
                norm_line = re.sub(r"(\d) (\d{1,3},)", r"\1\2", line)
                m = re.search(r"\$?\s*([\d,]{6,})", norm_line)
                if m:
                    return int(m.group(1).replace(",", ""))
        return None

    try:
        with pdfplumber.open(pdf_path) as pdf:
            result = _try_chinese(pdf)
            if result is None:
                result = _try_english(pdf)
            return result
    except Exception as e:
        logger.warning("PDF extraction failed [%s]: %s", pdf_path.name, e)
    return None


# Maps Chinese season text in E02 subjects to season number
_E02_SEASON_ZH = {
    "第一季": 1, "第1季": 1,
    "第二季": 2, "第2季": 2,
    "第三季": 3, "第3季": 3,
    "第四季": 4, "第4季": 4,
    "年度": 4,        # 年度 = annual = Q4
}

async def _search_e02_filings(stock_code: str, sdate: str = None) -> list[tuple[int, int]]:
    """Step 1: Search Chinese MOPS with PRO_ITEM=E02, lang=TW to discover available quarterly filing periods.
    Returns deduplicated list of (roc_year, season) pairs sorted newest-first."""
    data = {
        "step": "00", "RADIO_CM": "2", "TYPEK": "CO_MARKET",
        "CO_ID": stock_code, "PRO_ITEM": "E02", "SUBJECT": "",
        "SDATE": sdate or SDATE, "EDATE": EDATE, "lang": "TW",
    }
    try:
        async with httpx.AsyncClient(headers=_HEADERS, timeout=30, verify=False) as client:
            resp = await client.post(MOPSOV_SEARCH_URL, data=data)
            resp.raise_for_status()
            records = json.loads(resp.text.lstrip("﻿ \r\n\t")).get("data", [])
    except Exception as exc:
        logger.error("E02 search [%s]: %s", stock_code, exc)
        return []

    seen, results = set(), []
    for item in records:
        subject  = item.get("SUBJECT", "")
        date_str = item.get("CDATE", "")
        roc_year, season = _parse_e02_period(subject, date_str)
        if roc_year and season and (roc_year, season) not in seen:
            seen.add((roc_year, season))
            results.append((roc_year, season))
    results.sort(reverse=True)
    logger.info("E02 filings [%s]: %d periods found", stock_code, len(results))
    return results


def _parse_e02_period(subject: str, date_str: str) -> tuple[int | None, int | None]:
    """Extract (roc_year, season) from an E02 filing subject like '114年度第4季財務報告'."""
    ry_m = re.search(r"(\d{2,3})年", subject)
    roc_year = int(ry_m.group(1)) if ry_m else None

    season = None
    for label, s in _E02_SEASON_ZH.items():
        if label in subject:
            season = s
            break

    if roc_year and season:
        return roc_year, season

    # Fallback: derive period from filing date
    gm = re.match(r"(\d{4})/(\d{2})", date_str) if date_str else None
    if gm:
        return _filing_month_to_season(int(gm.group(1)), int(gm.group(2)))
    return None, None


def _pick_ai_link(links: list[tuple], season: int,
                  prefer_consolidated: bool = False) -> tuple | None:
    """Select the best AI link for the given season.
    prefer_consolidated=True  (Financial Holdings): AI1 > AIA only; never fall back to AI3.
    prefer_consolidated=False (all others):         prefer AI3; fall back to AI1 for Q4 only."""
    if prefer_consolidated:
        for rt in ("AI1", "AIA"):
            hit = [t for t in links if t[4] == season and t[5] == rt]
            if hit:
                return hit[0]
        return None
    else:
        ai3 = [t for t in links if t[4] == season and t[5] == "AI3"]
        if ai3:
            return ai3[0]
        if season == 4:
            ai1 = [t for t in links if t[4] == season and t[5] == "AI1"]
            if ai1:
                return ai1[0]
    return None


async def _get_ai3_for_period(
    client, stock_code: str, roc_year: int, season: int,
    target_display_code: str | None = None,
    prefer_consolidated: bool = False,
) -> tuple[str, str, str] | None:
    """Steps 2–4: Navigate t57sb01, find subsidiary if needed, locate AI PDF, get temp download URL.
    Returns (pdf_url, pdf_filename, subsidiary_internal_co_id) or None.
    prefer_consolidated=True → Financial Holdings: prefer AI1 (consolidated) for all seasons.
    prefer_consolidated=False → others: prefer AI3 (individual); AI1 fallback for Q4 only.
    """
    t57sb01_year = roc_year

    # Step 2: GET listing for the parent company (seamon empty = all seasons)
    await asyncio.sleep(1.5)
    r1 = await client.get(_twse_doc_url(stock_code, t57sb01_year, seamon=""))
    if r1.status_code != 200 or len(r1.text) < 500:
        return None
    soup1 = BeautifulSoup(r1.text, "html.parser")

    # Try to find AI links directly (company IS the reporting entity)
    ai_links = _extract_ai3_links(soup1)
    if ai_links:
        hit = _pick_ai_link(ai_links, season, prefer_consolidated=prefer_consolidated)
        if not hit:
            return None
        filename, kind, co_id = hit[0], hit[1], hit[2]
        pdf_url = await _get_twse_pdf_url(client, kind, co_id, filename)
        return (pdf_url, filename, co_id) if pdf_url else None

    # Financial Holdings only file under their own stock code — skip subsidiary navigation
    if prefer_consolidated:
        return None

    # Step 3: Find the relevant subsidiary's internal co_id
    sub_co_id, _ = _find_twse_subsidiary(soup1, target_display_code=target_display_code)
    if not sub_co_id:
        return None

    # Step 4: Navigate to subsidiary page (seamon empty) — shows IFRSs個別財報 links
    await asyncio.sleep(1.5)
    r2 = await client.get(_twse_doc_url(sub_co_id, t57sb01_year, seamon=""))
    if r2.status_code != 200 or len(r2.text) < 500:
        return None
    soup2 = BeautifulSoup(r2.text, "html.parser")

    ai_links = _extract_ai3_links(soup2)
    if not ai_links:
        return None
    hit = _pick_ai_link(ai_links, season, prefer_consolidated=prefer_consolidated)
    if not hit:
        return None
    filename, kind, co_id = hit[0], hit[1], hit[2]
    pdf_url = await _get_twse_pdf_url(client, kind, co_id, filename)
    return (pdf_url, filename, sub_co_id) if pdf_url else None


async def scrape_quarterly_reports(stock_code: str, roc_years: int = 2,
                                   subsidiary_name_en: str = "",
                                   subsidiaries: list | None = None,
                                   use_consolidated: bool = False) -> list[dict]:
    """Quarterly AUM extraction — downloads individual (AI3) PDFs for each listed subsidiary.
    subsidiaries: list of {"display_code": "28810012", "name_en": "Fubon Life Insurance"}.
    If omitted, falls back to single-subsidiary behaviour using f26_display_code / subsidiary_name_en.
    State key format: "{display_code}|{period_label}" to support multiple subsidiaries per period.
    """
    state_path = STATE_DIR / f"{stock_code}_balance_history.json"
    existing: dict[str, dict] = {}
    if state_path.exists():
        try:
            for r in json.loads(state_path.read_text(encoding="utf-8")):
                # Migrate old period-only keys to compound {display_code}|{period} keys
                sub_code = str(r.get("subsidiary_code") or "")
                disp = _FS_SUBSIDIARY_CODE_MAP.get(sub_code, r.get("stock_code", stock_code))
                existing[f"{disp}|{r['period']}"] = r
        except Exception:
            pass

    # Step 1: discover available periods from Chinese MOPS E02 search
    filing_periods = await _search_e02_filings(stock_code)
    if not filing_periods:
        logger.info("No E02 filings found for %s — no quarterly data to fetch", stock_code)
        _flush_balance_history(state_path, existing)
        return list(existing.values())

    entry = _WATCHLIST_MAP.get(stock_code, {})

    # Build target list: each entry is {"display_code": ..., "name_en": ...}
    if subsidiaries:
        targets = subsidiaries
        prefer_consolidated = False   # individual reports for each subsidiary
    else:
        tdc   = entry.get("f26_display_code") or None
        tname = subsidiary_name_en or entry.get("f26_name_en") or entry.get("name_en", stock_code)
        targets = [{"display_code": tdc, "name_en": tname}]
        prefer_consolidated = use_consolidated

    async with httpx.AsyncClient(headers=_TWSE_DOC_HEADERS, timeout=60,
                                  verify=False, follow_redirects=True) as client:
        for roc_year, season in filing_periods:
            period_label = f"{roc_year + 1911}/{_SEASON_LABEL.get(season, f'S{season}')}"
            season_str   = _SEASON_LABEL.get(season, f"S{season}")
            year_str     = str(roc_year + 1911)

            for target in targets:
                target_dc = target.get("display_code") or None
                sub_label = target.get("name_en") or entry.get("name_en", stock_code)
                ckey = f"{target_dc or stock_code}|{period_label}"
                prev = existing.get(ckey, {})
                sub_pdf_dir = PDF_DIR / (target_dc or stock_code)

                try:
                    # If re-trying a failed extraction and PDF is cached, skip download
                    prev_pdf = Path(prev["pdf_path"]) if prev.get("pdf_path") else None
                    if prev.get("extraction_failed") and prev_pdf and prev_pdf.exists():
                        pdf_dest        = prev_pdf
                        pdf_url         = prev.get("filing_url", "")
                        sub_co_id       = prev.get("subsidiary_code", stock_code)
                        is_consolidated = prev.get("is_consolidated", False)
                        logger.info("Re-trying extraction from cached PDF: %s", pdf_dest.name)
                    else:
                        # Steps 2–4: navigate t57sb01, find subsidiary, get AI PDF info
                        result = await _get_ai3_for_period(
                            client, stock_code, roc_year, season,
                            target_display_code=target_dc,
                            prefer_consolidated=prefer_consolidated,
                        )
                        if not result:
                            # For standalone companies (no specific subsidiary target) try MOPS iXBRL
                            if target_dc is None:
                                ixbrl = await _scrape_mops_ixbrl_quarterly(stock_code, roc_year, season)
                                if ixbrl:
                                    existing[ckey] = {
                                        "stock_code":                  stock_code,
                                        "subsidiary_code":             stock_code,
                                        "subsidiary_name_en":          sub_label,
                                        "period":                      ixbrl["period"],
                                        "roc_year":                    roc_year,
                                        "season":                      season,
                                        "is_consolidated":             ixbrl.get("is_consolidated", False),
                                        "currency":                    ixbrl.get("currency", "TWD"),
                                        "total_assets_raw":            ixbrl.get("total_assets_raw", ""),
                                        "total_assets_numeric":        ixbrl.get("total_assets_numeric"),
                                        "investment_property_raw":     ixbrl.get("investment_property_raw", ""),
                                        "investment_property_numeric": ixbrl.get("investment_property_numeric"),
                                        "source":                      "mops_ixbrl",
                                        "filing_url":                  ixbrl.get("filing_url", ""),
                                        "scraped_at":                  datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                                    }
                                    logger.info("iXBRL fallback stored [%s] %s", stock_code, period_label)
                            else:
                                logger.info("No AI PDF found [%s/%s] %s", stock_code, target_dc, period_label)
                            continue
                        pdf_url, orig_filename, sub_co_id = result

                        _orig_m = re.search(r"_(AI[13]|AIA)\.", orig_filename, re.I)
                        report_type     = _orig_m.group(1).upper() if _orig_m else "AI3"
                        is_consolidated = report_type in ("AI1", "AIA")
                        report_kind     = "Consolidated" if is_consolidated else "Individual"
                        pdf_name = f"{sub_label} {report_kind} Financial Report {season_str} {year_str}.pdf"
                        pdf_dest = sub_pdf_dir / pdf_name

                        sub_pdf_dir.mkdir(parents=True, exist_ok=True)
                        await asyncio.sleep(1.5)
                        pr = await client.get(pdf_url)
                        ct = pr.headers.get("content-type", "")
                        if pr.status_code != 200 or "pdf" not in ct:
                            logger.warning("PDF unavailable [%s/%s %s]: %d %s",
                                           stock_code, target_dc, period_label, pr.status_code, ct[:40])
                            continue
                        pdf_dest.write_bytes(pr.content)
                        logger.info("PDF saved: %s", pdf_dest.name)

                    # Step 6: extract 資産總計
                    total_assets = _extract_total_assets_from_pdf(pdf_dest)
                    if total_assets is None:
                        logger.warning("No 資産總計 in %s — flagged for manual check", pdf_dest.name)
                        existing[ckey] = {
                            "stock_code":                 stock_code,
                            "subsidiary_code":            sub_co_id,
                            "subsidiary_name_en":         sub_label,
                            "period":                     period_label,
                            "roc_year":                   roc_year,
                            "season":                     season,
                            "is_consolidated":            is_consolidated,
                            "currency":                   "TWD (thousands)",
                            "total_assets_raw":           "",
                            "total_assets_numeric":       None,
                            "investment_property_raw":    "",
                            "investment_property_numeric": None,
                            "pdf_path":                   str(pdf_dest),
                            "filing_url":                 pdf_url,
                            "scraped_at":                 datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                            "extraction_failed":          True,
                        }
                        continue

                    existing[ckey] = {
                        "stock_code":                 stock_code,
                        "subsidiary_code":            sub_co_id,
                        "subsidiary_name_en":         sub_label,
                        "period":                     period_label,
                        "roc_year":                   roc_year,
                        "season":                     season,
                        "is_consolidated":            is_consolidated,
                        "currency":                   "TWD (thousands)",
                        "total_assets_raw":           f"NT${total_assets:,}K",
                        "total_assets_numeric":       total_assets,
                        "investment_property_raw":    "",
                        "investment_property_numeric": None,
                        "pdf_path":                   str(pdf_dest),
                        "filing_url":                 pdf_url,
                        "scraped_at":                 datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                    }
                    logger.info("Quarterly BS [%s/%s] %s → NT$%s K",
                                stock_code, sub_label, period_label, f"{total_assets:,}")

                except Exception as e:
                    logger.warning("Quarterly reports [%s/%s] %s: %s",
                                   stock_code, target_dc, period_label, e)

    _flush_balance_history(state_path, existing)
    return sorted(existing.values(),
                  key=lambda r: (r.get("roc_year", 0), r.get("season", 0)), reverse=True)


def _flush_balance_history(state_path: Path, existing: dict) -> None:
    state_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.write_text(
        json.dumps(sorted(existing.values(),
                          key=lambda r: (r.get("roc_year", 0), r.get("season", 0)), reverse=True),
                   indent=2, ensure_ascii=False),
        encoding="utf-8"
    )


# ── EMOPS Company Profile + Consolidated Balance Sheet ────────────────────────

async def _post_emops(path: str, stock_code: str, extra: dict = None) -> str:
    """POST to an EMOPS endpoint, trying each TYPEK until a valid response is returned."""
    url = EMOPS_HOST + path
    base_data = {"co_id": stock_code, **(extra or {})}
    async with httpx.AsyncClient(headers=_EMOPS_PROFILE_HEADERS, timeout=30,
                                  follow_redirects=True, verify=False) as client:
        try:
            await client.get(f"{EMOPS_HOST}/server-java/t58query")
        except Exception:
            pass
        for typek in _TYPEK_OPTIONS:
            try:
                resp = await client.post(url, data={**base_data, "TYPEK": typek})
                resp.raise_for_status()
                html = resp.text
                if html and len(html) > 200 and "error" not in html[:200].lower():
                    logger.info("EMOPS POST %s [%s] TYPEK=%s → %d chars", path, stock_code, typek, len(html))
                    return html
            except Exception as exc:
                logger.warning("EMOPS POST failed [%s] TYPEK=%s: %s", stock_code, typek, exc)
    return ""


def _find_column_field(soup: BeautifulSoup, labels: list) -> str:
    """Find a field value by locating its column header label, then reading the cell below."""
    for label in labels:
        for cell in soup.find_all(["td", "th"],
                                   string=lambda t, l=label: t and l.lower() in t.lower()):
            row = cell.find_parent("tr")
            if not row:
                continue
            siblings = row.find_all(["td", "th"])
            try:
                col_idx = siblings.index(cell)
            except ValueError:
                continue
            next_row = row.find_next_sibling("tr")
            if next_row:
                next_cells = next_row.find_all(["td", "th"])
                if col_idx < len(next_cells):
                    val = next_cells[col_idx].get_text(strip=True)
                    if val and val.lower() != label.lower():
                        return val
    return ""


def _dedup_address(text: str) -> str:
    for marker in ["Taiwan", "R.O.C"]:
        idx = text.find(marker)
        if idx != -1:
            return text[:idx + len(marker)].strip(" ,")
    return text


_ADDR_LABEL_RE      = re.compile(
    r"(?i)^\s*(?:company\s+|registered\s+|business\s+|head\s+office\s+)?address\s*:?\s*$"
)
_NOT_ADDR_CONTENT_RE = re.compile(r"(?i)^https?://|^www\.|@")

def _find_emops_address(soup: BeautifulSoup) -> str:
    _ADDR_KW = ["Road", "Rd.", "St.", "Ave", "No.", "Blvd", "F.,", "路", "街", "號"]
    _CITY_KW = ["Taiwan", "Taipei", "Kaohsiung", "Taichung", "Tainan", "Hsinchu",
                "台北", "台中", "新北", "高雄", "桃園", "新竹", "台南"]

    def _valid(val):
        return (val and len(val) > 10
                and not _NOT_ADDR_CONTENT_RE.search(val)
                and not re.search(r"Address", val, re.I))

    # Strategy 1: label cell matching "Address", "Address:", "Company Address:", etc.
    # (but NOT "Web Address" / "Email Address" — excluded by _ADDR_LABEL_RE)
    for cell in soup.find_all(["td", "th"],
                               string=lambda t: t and _ADDR_LABEL_RE.match(t)):
        # (a) same-row next-sibling: inline label|value layout
        ns = cell.find_next_sibling(["td", "th"])
        if ns:
            val = ns.get_text(strip=True)
            if _valid(val):
                return _dedup_address(val)
        # (b) column-header layout: read same column position in next row
        row = cell.find_parent("tr")
        if not row:
            continue
        siblings = row.find_all(["td", "th"])
        try:
            col_idx = siblings.index(cell)
        except ValueError:
            continue
        next_row = row.find_next_sibling("tr")
        if next_row:
            cells = next_row.find_all(["td", "th"])
            if col_idx < len(cells):
                val = cells[col_idx].get_text(strip=True)
                if _valid(val):
                    return _dedup_address(val)
    # Strategy 2: colspan TDs with address keywords
    for td in soup.find_all("td", attrs={"colspan": True}):
        text = td.get_text(strip=True)
        if text and any(kw in text for kw in _ADDR_KW):
            return _dedup_address(text)
    # Strategy 3: any TD with address keywords + a city/country marker
    for td in soup.find_all("td"):
        text = td.get_text(strip=True)
        if (len(text) > 15
                and any(kw in text for kw in _ADDR_KW)
                and any(kw in text for kw in _CITY_KW)):
            return _dedup_address(text)
    return ""


async def scrape_emops_profile(stock_code: str) -> dict:
    html = await _post_emops("/server-java/t146sb05_e", stock_code, {"step": "0"})
    if not html:
        return {"stock_code": stock_code, "company_name_en": "", "address": "",
                "telephone": "", "web_address": ""}
    soup = BeautifulSoup(html, "lxml")
    return {
        "stock_code":      stock_code,
        "company_name_en": _find_column_field(soup, ["Company Name"]),
        "chairman":        _find_column_field(soup, ["Chairman"]),
        "general_manager": _find_column_field(soup, ["General Manager"]),
        "telephone":       _find_column_field(soup, ["Telephone"]),
        "web_address":     _find_column_field(soup, ["Web Address"]),
        "address":         _find_emops_address(soup),
    }


def _extract_consolidated_period(soup: BeautifulSoup) -> str:
    text = soup.get_text(" ")
    m = re.search(r"(20\d{2})[/-](\d{2})[/-](\d{2})", text)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
    m = re.search(r"民國\s*(\d{2,3})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", text)
    if m:
        return f"{int(m.group(1)) + 1911}/{int(m.group(2)):02d}/{int(m.group(3)):02d}"
    months = {"January":"01","February":"02","March":"03","April":"04","May":"05","June":"06",
              "July":"07","August":"08","September":"09","October":"10","November":"11","December":"12"}
    m = re.search(r"(January|February|March|April|May|June|July|August|September|October|"
                  r"November|December)\s+(\d{1,2}),?\s*(20\d{2})", text)
    if m:
        return f"{m.group(3)}/{months[m.group(1)]}/{int(m.group(2)):02d}"
    return ""


def _find_balance_value(soup: BeautifulSoup, labels: list) -> str:
    for label in labels:
        for cell in soup.find_all("td", string=lambda t, l=label:
                                   t and l.lower() in t.replace("　", "").strip().lower()):
            value_td = cell.find_next_sibling("td")
            if value_td:
                val = value_td.get_text(strip=True)
                if val:
                    return val
    return ""


async def _scrape_mops_ixbrl_quarterly(
    stock_code: str, roc_year: int, season: int
) -> dict | None:
    """Fallback quarterly balance sheet via MOPS t164sb01 iXBRL viewer.
    Used when no AI3/AI1 PDF exists on TWSE doc (e.g. companies that file iXBRL-only on MOPS).
    Tries consolidated (REPORT_ID=C) then individual (REPORT_ID=A)."""
    gregorian_year = roc_year + 1911
    period_label   = f"{gregorian_year}/{_SEASON_LABEL.get(season, f'S{season}')}"
    filing_url     = (f"https://mopsov.twse.com.tw/server-java/t164sb01"
                      f"?step=1&CO_ID={stock_code}&SYEAR={gregorian_year}"
                      f"&SSEASON={season}&REPORT_ID=C")
    for report_id in ("C", "A"):
        html = await _post_emops(
            "/server-java/t164sb01",
            stock_code,
            {"step": "1", "SYEAR": str(gregorian_year), "SSEASON": str(season),
             "REPORT_ID": report_id},
        )
        if not html:
            continue
        soup   = BeautifulSoup(html, "lxml")
        ta_raw = _find_balance_value(soup, ["Total assets", "Total Assets", "資產總計"])
        if not ta_raw:
            continue
        ip_raw   = _find_balance_value(soup, ["Investment property, net", "Investment property",
                                               "投資性不動產淨額", "投資性不動產"])
        period   = _extract_consolidated_period(soup) or period_label
        currency = "TWD (thousands)" if "千元" in soup.get_text() else "TWD"
        logger.info("iXBRL fallback [%s] %s → 資產總計=%s (REPORT_ID=%s)",
                    stock_code, period_label, ta_raw, report_id)
        return {
            "period":                      period,
            "is_consolidated":             report_id == "C",
            "currency":                    currency,
            "total_assets_raw":            ta_raw,
            "total_assets_numeric":        _parse_amount(ta_raw),
            "investment_property_raw":     ip_raw,
            "investment_property_numeric": _parse_amount(ip_raw),
            "filing_url":                  filing_url,
        }
    logger.info("iXBRL fallback [%s] %s → no data found", stock_code, period_label)
    return None


async def scrape_emops_balance_sheet(stock_code: str) -> dict:
    html = await _post_emops("/server-java/t164sb03_e", stock_code, {"step": "current"})
    base = {"stock_code": stock_code, "period": "", "currency": "TWD",
            "total_assets_raw": "", "total_assets_numeric": None,
            "investment_property_raw": "", "investment_property_numeric": None}
    if not html:
        return base
    soup = BeautifulSoup(html, "lxml")
    period   = _extract_consolidated_period(soup)
    currency = "TWD (thousands)" if "千元" in soup.get_text() else "TWD"
    ta_raw   = _find_balance_value(soup, ["Total assets", "Total Assets", "資產總計"])
    ip_raw   = _find_balance_value(soup, ["Investment property, net", "Investment property",
                                           "投資性不動產淨額", "投資性不動產"])
    return {
        "stock_code":                stock_code,
        "period":                    period,
        "currency":                  currency,
        "total_assets_raw":          ta_raw,
        "total_assets_numeric":      _parse_amount(ta_raw),
        "investment_property_raw":   ip_raw,
        "investment_property_numeric": _parse_amount(ip_raw),
    }


def _format_date(date_str: str) -> str:
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(date_str.strip(), fmt)
            return f"{dt.day} {dt.strftime('%B')} {dt.year}"
        except (ValueError, AttributeError):
            pass
    return date_str

# ── Change Detection ──────────────────────────────────────────────────────────

def _apply_date_status(records, new_since, date_field="announcement_date"):
    ns = new_since.replace("/", "-") if new_since else None
    for r in records:
        if ns:
            d = (r.get(date_field) or "").replace("/", "-")[:10]
            r["status"] = "NEW" if d >= ns else "HISTORICAL"
        else:
            r["status"] = "HISTORICAL"
    return records

def detect_changes(records, state_path, key_fields):
    stored = _load_json(state_path) or {}
    updated = dict(stored)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for record in records:
        key = "|".join(str(record.get(f, "")) for f in key_fields)
        h = hashlib.sha256(json.dumps(
            {k: v for k, v in record.items() if k not in {"scraped_at", "status", "hash"}},
            sort_keys=True, ensure_ascii=False).encode()).hexdigest()
        record["hash"] = h
        record["scraped_at"] = now
        if key not in stored:           record["status"] = "NEW"
        elif stored[key]["hash"] != h:  record["status"] = "CHANGED"
        else:                           record["status"] = "HISTORICAL"
        updated[key] = {"hash": h, "last_seen": now}
    state_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.write_text(json.dumps(updated, indent=2, ensure_ascii=False), encoding="utf-8")
    return records

def archive(stock_code, category, records):
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    path = ARCHIVE_DIR / f"{stock_code}_{category}_{ts}.json"
    path.write_text(json.dumps({"stock_code": stock_code, "category": category,
                                "records": records}, indent=2, ensure_ascii=False), encoding="utf-8")

def _load_json(path):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return None

_MOJIBAKE_RE = re.compile(chr(0xfffd) + '[@AB]')

def _clean_address(addr: str) -> str:
    if not addr:
        return addr
    addr = _MOJIBAKE_RE.sub(', ', addr)
    addr = addr.replace('、', ', ').replace('，', ', ')
    # Remove duplicate commas
    addr = re.sub(r',\s*,', ', ', addr)
    # Ensure exactly one space after every comma
    addr = re.sub(r',\s*', ', ', addr)
    # Add period+space after road abbreviations directly followed by a letter/digit with no space
    # e.g. "RdTaipei" → "Rd. Taipei", "Rd.Taipei" → "Rd. Taipei"
    addr = re.sub(r'\b(Rd|St|Ave|Blvd|Ln|Dr|Sec)\.?([A-Z\d])', r'\1. \2', addr)
    # Ensure space after No./Sec./F. before a digit: "No.100" → "No. 100"
    addr = re.sub(r'\b(No|Sec|Fl?)\.(\S)', r'\1. \2', addr)
    # Collapse multiple spaces
    addr = re.sub(r'\s+', ' ', addr).strip()
    # Normalise to title case only if the string is mostly uppercase
    if addr == addr.upper():
        addr = addr.title()
    return addr

def _clean_web_address(url: str) -> str:
    if not url:
        return url
    # Strip mojibake artifacts entirely from URLs
    url = _MOJIBAKE_RE.sub('', url).rstrip('@').strip()
    return url

_CO_SUFFIX_RE = re.compile(
    r'(?:[,.\s]|\s)*\b('
    r'co\.,?\s*ltd\.?|co\.,?\s*limited|company\s+limited|company\s+ltd\.?|'
    r'corp\.,?\s*ltd\.?|holdings?\s+company\s+limited|'
    r'incorporated|\blimited\b|ltd\.?|corp\.?|\bcorporation\b|\binc\.?|co\.?'
    r')\s*$',
    re.IGNORECASE
)

def _clean_company_name(name: str) -> str:
    if not name:
        return name
    name = name.strip()
    prev = None
    while prev != name:
        prev = name
        name = _CO_SUFFIX_RE.sub('', name).strip().rstrip(',').strip()
    if name == name.upper():
        name = name.title()
    return name

def _format_tw_phone(phone: str) -> str:
    """Normalize any Taiwan phone format to +886-X-XXXX-XXXX (landline) or +886-9XX-XXX-XXX (mobile).
    Always outputs exactly 3 dashes regardless of input separators."""
    if not phone:
        return phone
    digits = re.sub(r'\D', '', phone)
    if not digits:
        return phone
    # Strip country code then leading zero — always work from bare local digits
    if digits.startswith('886'):
        digits = digits[3:]
    if digits.startswith('0'):
        digits = digits[1:]
    n = len(digits)
    if n == 9:
        if digits[0] == '9':
            # Mobile (09X-XXX-XXX): +886-9XX-XXX-XXX
            return f'+886-{digits[:3]}-{digits[3:6]}-{digits[6:]}'
        else:
            # Landline (0X-XXXX-XXXX): +886-X-XXXX-XXXX
            return f'+886-{digits[0]}-{digits[1:5]}-{digits[5:]}'
    if n == 8:
        # 8-digit Taipei local without area code: +886-2-XXXX-XXXX
        return f'+886-2-{digits[:4]}-{digits[4:]}'
    return phone

# ── Excel Output ──────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="1F3864")
_NEW_FILL  = PatternFill("solid", fgColor="FFD28A")   # orange
_CHG_FILL  = PatternFill("solid", fgColor="FFEB9C")   # yellow
_HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

def _save_run_to_history(fund_commitments, people_moves, emops_data, since, new_since) -> list:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    history_path = OUTPUT_DIR / "run_history.json"
    history = []
    if history_path.exists():
        try:
            history = json.loads(history_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    def slim(r, keys):
        return {k: (r.get(k) or "") for k in keys}

    fc_keys = ["stock_code","announcement_date","fund_name","fund_type","commitment_date",
               "commitment_amount_raw","headline","key_events","internal_notes",
               "bs_date","status","scraped_at","url","fx_url"]
    pm_keys = ["stock_code","announcement_date","role_type","role_title","new_holder","previous_holder",
               "change_type","reason","effective_date","narrative_en","key_events","internal_notes",
               "bs_date","status","scraped_at","url"]
    em_keys = ["stock_code","name_en","company_type","company_name_en","address","telephone",
               "web_address","period","currency","total_assets_raw","inv_property_raw",
               "profile_status","changed_fields","scraped_at","no_filing_data"]

    def slim_em(r):
        d = {k: (r.get(k) or "") for k in em_keys}
        d["changed_fields"] = r.get("changed_fields") or []
        return d

    run = {
        "run_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "since": since or "",
        "new_since": new_since or "",
        "n_new_fc":  sum(1 for r in fund_commitments if r.get("status") == "NEW"),
        "n_his_fc":  sum(1 for r in fund_commitments if r.get("status") == "HISTORICAL"),
        "n_new_pm":  sum(1 for r in people_moves    if r.get("status") == "NEW"),
        "n_chg_pm":  sum(1 for r in people_moves    if r.get("status") == "CHANGED"),
        "n_chg_em":  sum(1 for r in (emops_data or []) if r.get("profile_status") == "CHANGED"),
        "funds":  [slim(r, fc_keys) for r in fund_commitments],
        "people": [slim(r, pm_keys) for r in people_moves],
        "emops":  [slim_em(r) for r in (emops_data or [])],
    }
    history.insert(0, run)
    history_path.write_text(json.dumps(history, indent=2, ensure_ascii=False), encoding="utf-8")
    return history


def write_html_report(fund_commitments, people_moves, emops_data=None, balance_history=None, since=None, new_since=None):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    history = _save_run_to_history(fund_commitments, people_moves, emops_data, since, new_since)
    history_json = json.dumps(history, ensure_ascii=False)
    balance_history_json = json.dumps(balance_history or [], ensure_ascii=False)
    run_ts = history[0]["run_date"]

    run_options = "\n".join(
        f'<option value="{i}">{r["run_date"].split(" ")[0]}  —  {r["n_new_fc"]} new FC · {r["n_new_pm"]} new PM · {r.get("n_chg_em",0)} EMOPS changes</option>'
        for i, r in enumerate(history)
    )

    companies_js = "{" + ",".join(
        f'"{w["stock_code"]}":"{w["name_en"]}"' for w in WATCHLIST
    ) + "," + ",".join(
        f'"{s["stock_code"]}":"{s["company_name_en"]}"' for s in _SUBSIDIARY_STUBS
    ) + "}"
    firm_ids_js   = "{" + ",".join(f'"{w["stock_code"]}":"{w.get("firm_id","")}"'      for w in WATCHLIST) + "}"
    firm_types_js = "{" + ",".join(f'"{w["stock_code"]}":"{w.get("company_type","")}"' for w in WATCHLIST) + "}"
    firm_urls_js  = "{" + ",".join(f'"{w["stock_code"]}":"{w.get("firm_url","")}"'     for w in WATCHLIST) + "}"

    # JS is a plain string (not f-string) so template literals work without escaping
    js = (
        f"<script>\nconst COMPANIES={companies_js};\nconst FIRM_IDS={firm_ids_js};\nconst FIRM_TYPES={firm_types_js};\nconst FIRM_URLS={firm_urls_js};\nconst _API_PORT={API_PORT};\nconst HISTORY = "
        + history_json
        + f";\nconst BALANCE_HISTORY = "
        + balance_history_json
        + r"""
;
function badge(st){const m={NEW:'new',HISTORICAL:'his',CHANGED:'chg',UNCHANGED:''};const c=m[st]||'';return c?`<span class="badge ${c}">${st}</span>`:(st||'');}
function fv(val,changed){return changed?`<span class="field-chg" title="Changed since last run">${val||'—'}</span>`:(val||'—');}
function fmtHeadline(hl){if(!hl)return '';const parts=hl.split('\n\n');if(parts.length<2)return hl;return `<strong>${parts[0]}</strong><br><span style="font-style:italic">${parts.slice(1).join('<br>')}</span>`;}
function firmCell(code,name){const u=FIRM_URLS[code];return u?`<a href="${u}" target="_blank">${name}</a>`:name;}
// Review-state sync: uses shared API when reachable, falls back to localStorage per-browser.
// API server is started by calling mopsov.start_api_server() from your app wrapper.
(function(){
  var _h='';
  try{_h=window.parent.location.hostname;}catch(e){}
  if(!_h)try{_h=window.location.hostname;}catch(e){}
  if(!_h)_h='localhost';
  var _api=window.location.protocol+'//'+_h+':'+_API_PORT+'/api/checks';
  var _mem={};var _useApi=false;
  window.chkGet=function(key){return _mem[key]||null;};
  window.chkSet=function(key,state,name){
    var d={state:state,name:name,ts:new Date().toISOString()};
    _mem[key]=d;
    if(_useApi){
      fetch(_api,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({[key]:d})}).catch(function(){});
    }else{try{localStorage.setItem(key,JSON.stringify(d));}catch(e){}}
  };
  function _applyAll(){
    document.querySelectorAll('.chk-wrap').forEach(function(wrap){
      var btn=wrap.querySelector('.chk-btn');if(!btn)return;
      var m=(btn.getAttribute('onclick')||'').match(/toggleCheck\(this,'([^']+)'\)/);if(!m)return;
      var key=m[1];var d=_mem[key]||{};var st=d.state||'pending';
      var lbl={pending:'Pending',checked:'Checked ✓',irrelevant:'Irrelevant'};
      var cls={pending:'chk-pend',checked:'chk-done',irrelevant:'chk-irrel'};
      btn.textContent=lbl[st]||'Pending';btn.className='chk-btn '+(cls[st]||'chk-pend');
      var inp=wrap.querySelector('.chk-name');if(inp&&d.name)inp.value=d.name;
      var tsEl=wrap.querySelector('.chk-ts');if(tsEl&&d.ts&&typeof fmtTs==='function')tsEl.textContent=fmtTs(d.ts);
    });
  }
  function _sync(){
    fetch(_api).then(function(r){return r.json();}).then(function(data){
      _useApi=true;Object.assign(_mem,data);_applyAll();
    }).catch(function(){
      if(!_useApi){
        try{for(var i=0;i<localStorage.length;i++){var k=localStorage.key(i);
          try{var v=JSON.parse(localStorage.getItem(k));if(v&&v.state)_mem[k]=v;}catch(e){}}
          _applyAll();
        }catch(e){}
      }
    });
  }
  document.addEventListener('DOMContentLoaded',function(){setTimeout(_sync,250);});
  setInterval(_sync,30000);
})();
function fmtTs(iso){if(!iso)return '';try{const d=new Date(iso);return d.toLocaleDateString('en-GB',{day:'2-digit',month:'short',year:'numeric'})+' '+d.toLocaleTimeString('en-GB',{hour:'2-digit',minute:'2-digit'});}catch{return '';}}
function copyNotes(btn){
  const text=decodeURIComponent(escape(atob(btn.dataset.n||'')));
  (navigator.clipboard?navigator.clipboard.writeText(text):Promise.reject()).catch(()=>{
    const ta=document.createElement('textarea');ta.value=text;ta.style.position='fixed';ta.style.opacity='0';
    document.body.appendChild(ta);ta.select();document.execCommand('copy');document.body.removeChild(ta);
  }).then(()=>{}).catch(()=>{});
  btn.textContent='Copied!';btn.classList.add('copied');
  setTimeout(()=>{btn.textContent='Copy notes';btn.classList.remove('copied');},2000);
}
function chkBtn(key){
  const d=chkGet(key)||{};const st=d.state||'pending';
  const lbl={pending:'Pending',checked:'Checked ✓',irrelevant:'Irrelevant'};
  const cls={pending:'chk-pend',checked:'chk-done',irrelevant:'chk-irrel'};
  return `<span class="chk-wrap"><button class="chk-btn ${cls[st]||'chk-pend'}" onclick="toggleCheck(this,'${key}')">${lbl[st]||'Pending'}</button><div class="chk-ts">${fmtTs(d.ts)}</div><input class="chk-name" placeholder="Reviewer" value="${d.name||''}" onchange="saveName(this,'${key}')" onclick="event.stopPropagation()"></span>`;
}
function toggleCheck(btn,key){
  const inp=btn.closest('.chk-wrap').querySelector('.chk-name');
  if(!inp.value.trim()){inp.style.outline='2px solid #e74c3c';inp.focus();return;}
  inp.style.outline='';
  const d=chkGet(key)||{};const cycle=['pending','checked','irrelevant'];
  const next=cycle[(cycle.indexOf(d.state||'pending')+1)%3];
  chkSet(key,next,inp.value.trim());
  const lbl={pending:'Pending',checked:'Checked ✓',irrelevant:'Irrelevant'};
  const cls={pending:'chk-pend',checked:'chk-done',irrelevant:'chk-irrel'};
  btn.textContent=lbl[next];btn.className='chk-btn '+cls[next];
  btn.nextElementSibling.textContent=fmtTs(new Date().toISOString());
}
function saveName(inp,key){const d=chkGet(key)||{state:'pending'};chkSet(key,d.state,inp.value);}

function populateSel(sel,opts){
  const cur=sel.value;sel.innerHTML='<option value="">All</option>';
  opts.forEach(([v,l])=>{const o=document.createElement('option');o.value=v;o.textContent=l;sel.appendChild(o);});
  if([...sel.options].some(o=>o.value===cur))sel.value=cur;
}

function renderRun(idx){
  const run=HISTORY[idx];
  document.querySelectorAll('.filters input[type=text]').forEach(i=>i.value='');
  document.getElementById('run-info').textContent=`Data from: ${run.since||'2 years'} · baseline: ${run.new_since||'state file'}`;
  document.getElementById('m-new-fc').textContent=run.n_new_fc;
  document.getElementById('m-his-fc').textContent=run.n_his_fc;
  document.getElementById('m-new-pm').textContent=run.n_new_pm;
  // Populate company dropdowns from run data
  const nameMap=Object.assign({},COMPANIES);BALANCE_HISTORY.forEach(r=>{if(r.name_en)nameMap[r.stock_code]=r.name_en;});
  const allCodes=[...new Set([...run.funds.map(r=>r.stock_code),...(run.people||[]).map(r=>r.stock_code),...(run.emops||[]).map(r=>r.stock_code)])].sort();
  const codeOpts=allCodes.map(c=>[c,nameMap[c]?`${c} — ${nameMap[c]}`:c]);
  document.querySelectorAll('.co-filter:not(#fs .co-filter)').forEach(s=>populateSel(s,codeOpts));
  // Populate fund type dropdown
  const ftypes=[...new Set(run.funds.map(r=>r.fund_type||'').filter(Boolean))].sort();
  populateSel(document.getElementById('fc-ft-sel'),ftypes.map(t=>[t.toLowerCase(),t]));
  renderEM(run.emops||[]);renderFS(BALANCE_HISTORY);renderFC(run.funds||[]);renderPM(run.people||[]);
}
function renderEM(rows){
  document.querySelector('#em-table tbody').innerHTML=rows.map(r=>{
    if(r.no_filing_data){
      return `<tr data-co="${r.stock_code}" data-date=""><td>${r.stock_code}</td><td>${FIRM_IDS[r.stock_code]||''}</td><td>${firmCell(r.stock_code,r.company_name_en)}</td><td colspan="7" style="color:#888;font-style:italic;">Information not available on filings — please refer to firm website</td><td></td></tr>`;
    }
    const webUrl=r.web_address?(r.web_address.match(/^https?:\/\//)?r.web_address:'https://'+r.web_address):'';
    const webCell=webUrl?`<a href="${webUrl}" target="_blank">${r.web_address}</a>`:(r.web_address||'—');
    const ck=`chk_em_${r.stock_code}`;
    return `<tr data-co="${r.stock_code}" data-date="${normDate(r.scraped_at||'')}"><td>${r.stock_code}</td><td>${FIRM_IDS[r.stock_code]||''}</td><td>${firmCell(r.stock_code,r.name_en||r.company_name_en||'—')}</td><td>${FIRM_TYPES[r.stock_code]||''}</td><td>${r.period||'—'}</td><td>${r.telephone||'—'}</td><td>${webCell}</td><td>${r.address||'—'}</td><td>${fmtScraped(r.scraped_at)}</td><td>${chkBtn(ck)}</td></tr>`;
  }).join('');
  document.getElementById('em-count').textContent=rows.length+' companies';
}
function renderFS(rows){
  const fsCos=[...new Map(rows.map(r=>[r.stock_code,r.name_en||COMPANIES[r.stock_code]||r.stock_code])).entries()].sort((a,b)=>a[0].localeCompare(b[0]));
  populateSel(document.querySelector('#fs .filters select'),fsCos.map(([c,n])=>[c,`${c} — ${n}`]));
  document.querySelector('#fs-table tbody').innerHTML=rows.map(r=>{
    const base=r.name_en||COMPANIES[r.stock_code]||r.stock_code;
    const co=r.delisted?`${base} <span style="color:#c0392b;font-weight:600">(Delisted)</span>`:base;
    const src=r.stub_only?'':r.source==='quarterly'?'PDFs - Unconsolidated FS':r.source==='quarterly_consolidated'?'PDFs - Consolidated FS':'Balance Sheet';
    const ck=`chk_fs_${r.stock_code}_${(r.period||'').replace(/\W+/g,'_')}`;
    const filing=r.source==='annual'&&r.filing_url?`<a href="${r.filing_url}" target="_blank">View ↗</a>`:(r.pdf_filename||'—');
    const taCell=r.extraction_failed
      ?`<span style="color:#e67e22;font-weight:600" title="PDF downloaded but text could not be extracted (likely image-based). Manual review required.">⚠ Manual check needed</span>`
      :(r.total_assets_raw||'—');
    const rowCls=r.extraction_failed?'style="background:#fff8f0"':'';
    const rawNotes=r.internal_notes||'';
    const notes=rawNotes?`<button class="copy-btn" data-n="${btoa(unescape(encodeURIComponent(rawNotes)))}" onclick="copyNotes(this)">Copy notes</button>${rawNotes.replace(/\n/g,'<br>')}`:''
    return `<tr ${rowCls} data-co="${r.stock_code}" data-period="${(r.period||'').toLowerCase()}"><td>${r.stock_code}</td><td>${FIRM_IDS[r.stock_code]||''}</td><td>${firmCell(r.stock_code,co)}</td><td>${FIRM_TYPES[r.stock_code]||''}</td><td>${r.period||'—'}</td><td>${src}</td><td>${taCell}</td><td>${r.investment_property_raw||'—'}</td><td>${fmtScraped(r.scraped_at)}</td><td>${filing}</td><td class="headline">${notes}</td><td>${chkBtn(ck)}</td></tr>`;
  }).join('');
  document.getElementById('fs-count').textContent=rows.length+' records';
}
function renderFC(rows){
  document.querySelector('#fc-table tbody').innerHTML=rows.map(r=>{
    const st=r.status||'';const yr=(r.announcement_date||'').slice(0,4);
    const nm=r.url?`<a href="${r.url}" target="_blank">${r.fund_name}</a>`:(r.fund_name||'');
    const amt=r.commitment_amount_raw?(r.fx_url?`<a href="${r.fx_url}" target="_blank">${r.commitment_amount_raw}</a>`:r.commitment_amount_raw):'—';
    const ck=`chk_fc_${r.stock_code}_${(r.fund_name||'').replace(/\W+/g,'_')}_${r.commitment_date}`;
    const firm=COMPANIES[r.stock_code]||r.stock_code;
    const rawNotes=r.internal_notes||'';
    const notes=rawNotes?`<button class="copy-btn" data-n="${btoa(unescape(encodeURIComponent(rawNotes)))}" onclick="copyNotes(this)">Copy notes</button>${rawNotes.replace(/\n/g,'<br>')}`:''
    return `<tr class="row-${st.toLowerCase()}" data-co="${r.stock_code}" data-ft="${(r.fund_type||'').toLowerCase()}" data-st="${st.toLowerCase()}" data-yr="${yr}" data-date="${normDate(r.announcement_date||'')}"><td>${r.stock_code}</td><td>${FIRM_IDS[r.stock_code]||''}</td><td>${firmCell(r.stock_code,firm)}</td><td>${FIRM_TYPES[r.stock_code]||''}</td><td>${r.announcement_date}</td><td>${nm}</td><td>${r.fund_type||'—'}</td><td>${r.commitment_date}</td><td>${amt}</td><td class="headline">${fmtHeadline(r.headline||'')}</td><td class="headline">${r.key_events||''}</td><td>${r.bs_date}</td><td>${fmtScraped(r.scraped_at)}</td><td class="headline">${notes}</td><td>${chkBtn(ck)}</td></tr>`;
  }).join('');
  document.getElementById('fc-count').textContent=rows.length+' records';
}
function renderPM(rows){
  document.querySelector('#pm-table tbody').innerHTML=rows.map(r=>{
    const st=r.status||'';const yr=(r.announcement_date||'').slice(0,4);
    const role=r.role_title||r.role_type||'';
    const lnk=r.url?`<a href="${r.url}" target="_blank">View</a>`:'';
    const ck=`chk_pm_${r.stock_code}_${r.announcement_date}_${(r.new_holder||'').replace(/\W+/g,'_')}`;
    const firmNm=COMPANIES[r.stock_code]||r.stock_code;
    const rawNotes=r.internal_notes||'';
    const notes=rawNotes?`<button class="copy-btn" data-n="${btoa(unescape(encodeURIComponent(rawNotes)))}" onclick="copyNotes(this)">Copy notes</button>${rawNotes.replace(/\n/g,'<br>')}`:''
    return `<tr class="row-${st.toLowerCase()}" data-co="${r.stock_code}" data-st="${st.toLowerCase()}" data-yr="${yr}" data-date="${normDate(r.announcement_date||'')}"><td>${r.stock_code}</td><td>${FIRM_IDS[r.stock_code]||''}</td><td>${firmCell(r.stock_code,firmNm)}</td><td>${FIRM_TYPES[r.stock_code]||''}</td><td>${r.announcement_date}</td><td>${role||'—'}</td><td>${r.new_holder||'—'}</td><td>${r.previous_holder||'—'}</td><td>${r.effective_date}</td><td class="headline">${fmtHeadline(r.narrative_en||'')}</td><td class="headline">${r.key_events||''}</td><td>${r.bs_date||'—'}</td><td>${lnk}</td><td>${fmtScraped(r.scraped_at)}</td><td class="headline">${notes}</td><td>${chkBtn(ck)}</td></tr>`;
  }).join('');
  document.getElementById('pm-count').textContent=rows.length+' records';
}
function showTab(id,el){document.querySelectorAll('.panel').forEach(p=>p.classList.remove('active'));document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));document.getElementById(id).classList.add('active');el.classList.add('active');}
function normDate(d){return d?d.replace(/\//g,'-'):'';}
function inDateRange(rowDate,from,to){const d=normDate(rowDate);return(!from||d>=from)&&(!to||d<=to);}
function fmtScraped(s){if(!s)return '—';return s.slice(0,10).replace(/-/g,'/');}

function filterEM(){
  const co=document.querySelector('#em .filters select').value.toLowerCase();
  const dates=[...document.querySelectorAll('#em .filters input[type=date]')].map(e=>e.value);
  const q=document.querySelector('#em .filters input[type=text]').value.toLowerCase();
  const [df,dt]=dates;
  let v=0;document.querySelectorAll('#em-table tbody tr').forEach(r=>{
    const s=(!co||r.dataset.co===co)&&inDateRange(r.dataset.date,df,dt)&&(!q||r.textContent.toLowerCase().includes(q));
    r.style.display=s?'':'none';if(s)v++;});
  document.getElementById('em-count').textContent=v+' companies';
}
function filterFS(){
  const co=document.querySelector('#fs .filters select').value.toLowerCase();
  const q=document.querySelector('#fs .filters input[type=text]').value.toLowerCase();
  let v=0;document.querySelectorAll('#fs-table tbody tr').forEach(r=>{
    const s=(!co||r.dataset.co===co)&&(!q||r.textContent.toLowerCase().includes(q));
    r.style.display=s?'':'none';if(s)v++;});
  document.getElementById('fs-count').textContent=v+' records';
}
function filterFC(){
  const sel=[...document.querySelectorAll('#fc .filters select')].map(e=>e.value.toLowerCase());
  const dates=[...document.querySelectorAll('#fc .filters input[type=date]')].map(e=>e.value);
  const q=document.querySelector('#fc .filters input[type=text]').value.toLowerCase();
  const [co,ft]=[sel[0],sel[1]];const [df,dt]=dates;
  let v=0;document.querySelectorAll('#fc-table tbody tr').forEach(r=>{
    const s=(!co||r.dataset.co===co)&&(!ft||r.dataset.ft.includes(ft))&&inDateRange(r.dataset.date,df,dt)&&(!q||r.textContent.toLowerCase().includes(q));
    r.style.display=s?'':'none';if(s)v++;});
  document.getElementById('fc-count').textContent=v+' records';
}
function filterPM(){
  const sel=[...document.querySelectorAll('#pm .filters select')].map(e=>e.value.toLowerCase());
  const dates=[...document.querySelectorAll('#pm .filters input[type=date]')].map(e=>e.value);
  const q=document.querySelector('#pm .filters input[type=text]').value.toLowerCase();
  const [co,st]=[sel[0],sel[1]];const [df,dt]=dates;
  let v=0;document.querySelectorAll('#pm-table tbody tr').forEach(r=>{
    const s=(!co||r.dataset.co===co)&&(!st||r.dataset.st===st)&&inDateRange(r.dataset.date,df,dt)&&(!q||r.textContent.toLowerCase().includes(q));
    r.style.display=s?'':'none';if(s)v++;});
  document.getElementById('pm-count').textContent=v+' records';
}
const _sortState={};
function sortTable(tid,col){
  const t=document.getElementById(tid);
  const st=_sortState[tid]||{col:-1,asc:true};
  const asc=st.col===col?!st.asc:true;
  _sortState[tid]={col,asc};
  t.querySelectorAll('thead th').forEach((th,i)=>{if(th.classList.contains('sortable'))th.dataset.sort=i===col?(asc?'asc':'desc'):'';});
  const rows=[...t.querySelectorAll('tbody tr')];
  rows.sort((a,b)=>{
    const av=(a.cells[col]?.textContent||'').trim();
    const bv=(b.cells[col]?.textContent||'').trim();
    const an=parseFloat(av.replace(/[^0-9.-]/g,''));
    const bn=parseFloat(bv.replace(/[^0-9.-]/g,''));
    if(!isNaN(an)&&!isNaN(bn))return asc?an-bn:bn-an;
    return asc?av.localeCompare(bv):bv.localeCompare(av);
  });
  const tb=t.querySelector('tbody');rows.forEach(r=>tb.appendChild(r));
}
renderRun(0);
</script>"""
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>MOPs</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{font-family:'Segoe UI',sans-serif;background:#f5f6fa;color:#1a1a2e;}}
  header{{background:#FF6633;color:#fff;padding:14px 32px;display:flex;align-items:center;gap:20px;flex-wrap:wrap;}}
  header h1{{font-size:1.25rem;font-weight:700;flex-shrink:0;}}
  .run-sel{{display:flex;align-items:center;gap:8px;}}
  .run-sel label{{font-size:.8rem;opacity:.75;white-space:nowrap;}}
  .run-sel select{{padding:4px 10px;border-radius:6px;border:none;font-size:.82rem;background:rgba(0,0,0,.15);color:#fff;cursor:pointer;max-width:360px;}}
  .run-sel select option{{background:#e05520;}}
  #run-info{{font-size:.78rem;opacity:.65;margin-left:auto;}}
  .metrics{{display:flex;gap:16px;padding:20px 32px;}}
  .metric{{background:#fff;border-radius:8px;padding:14px 20px;flex:1;box-shadow:0 1px 4px rgba(0,0,0,.08);}}
  .metric .num{{font-size:1.9rem;font-weight:700;color:#FF6633;}}
  .metric .lbl{{font-size:.78rem;color:#666;margin-top:4px;}}
  .tabs{{display:flex;padding:0 32px;border-bottom:2px solid #ddd;margin-top:4px;}}
  .tab{{padding:10px 22px;cursor:pointer;font-weight:600;font-size:.88rem;color:#666;border-bottom:3px solid transparent;margin-bottom:-2px;}}
  .tab.active{{color:#FF6633;border-bottom-color:#FF6633;}}
  .panel{{display:none;padding:20px 32px;}}
  .panel.active{{display:block;}}
  .filters{{display:flex;gap:10px;margin-bottom:14px;flex-wrap:wrap;align-items:center;}}
  .filters select,.filters input{{padding:5px 10px;border:1px solid #ddd;border-radius:6px;font-size:.83rem;background:#fff;}}
  .filters label{{font-size:.78rem;color:#666;font-weight:600;}}
  table{{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.08);font-size:.81rem;}}
  th{{background:#FF6633;color:#fff;padding:9px 11px;text-align:left;font-weight:600;white-space:nowrap;}}
  th.sortable{{cursor:pointer;user-select:none;}}
  th.sortable:hover{{background:#e05520;}}
  th.sortable::after{{content:' ⇅';opacity:.35;font-size:.7em;}}
  th.sortable[data-sort=asc]::after{{content:' ▲';opacity:1;}}
  th.sortable[data-sort=desc]::after{{content:' ▼';opacity:1;}}
  td{{padding:8px 11px;border-bottom:1px solid #f0f0f0;vertical-align:top;}}
  tr:last-child td{{border-bottom:none;}}
  tr:hover td{{background:rgba(0,0,0,.02);}}
  .row-new td{{background:#FFE0B2;}}.row-changed td{{background:#fffbf0;}}
  .headline{{max-width:380px;font-style:italic;color:#444;}}
  .badge{{padding:2px 7px;border-radius:4px;font-size:.73rem;font-weight:700;white-space:nowrap;}}
  .badge.new{{background:#C6EFCE;color:#276221;}}.badge.his{{background:#FFC7CE;color:#9C0006;}}.badge.chg{{background:#FFEB9C;color:#9C6500;}}
  .field-chg{{background:#FFEB9C;padding:1px 5px;border-radius:3px;font-weight:600;}}
  .cf-list{{font-size:.75rem;color:#9C6500;font-style:italic;}}
  .chk-wrap{{display:flex;flex-direction:column;gap:2px;min-width:88px;}}
  .chk-btn{{padding:3px 8px;border:none;border-radius:4px;font-size:.75rem;font-weight:700;cursor:pointer;white-space:nowrap;width:100%;}}
  .chk-pend{{background:#FFD700;color:#5a4000;}}.chk-done{{background:#C6EFCE;color:#276221;}}.chk-irrel{{background:#D3D3D3;color:#555;}}
  .chk-ts{{font-size:.65rem;color:#888;white-space:nowrap;}}
  .chk-name{{width:100%;padding:2px 4px;border:1px solid #ddd;border-radius:3px;font-size:.72rem;box-sizing:border-box;}}
  .copy-btn{{padding:2px 7px;border:1px solid #ccc;border-radius:3px;font-size:.68rem;cursor:pointer;background:#f5f5f5;color:#555;white-space:nowrap;margin-bottom:4px;display:block;}}
  .copy-btn:hover{{background:#e8e8e8;}}.copy-btn.copied{{background:#C6EFCE;color:#276221;border-color:#9fd49f;}}
  a{{color:#FF6633;}}.count{{font-size:.78rem;color:#666;margin-bottom:8px;}}
</style>
</head>
<body>
<header>
  <h1>MOPs</h1>
  <div class="run-sel">
    <label>Run:</label>
    <select onchange="renderRun(+this.value)">
{run_options}
    </select>
  </div>
  <span id="run-info"></span>
</header>
<div class="metrics">
  <div class="metric"><div class="num" id="m-new-fc">—</div><div class="lbl">New Fund Commitments</div></div>
  <div class="metric"><div class="num" id="m-his-fc">—</div><div class="lbl">Historical Commitments</div></div>
  <div class="metric"><div class="num" id="m-new-pm">—</div><div class="lbl">New People Moves</div></div>
</div>
<div class="tabs">
  <div class="tab active" onclick="showTab('em',this)">Company Profiles</div>
  <div class="tab"        onclick="showTab('fs',this)">Financial Statements</div>
  <div class="tab"        onclick="showTab('fc',this)">Fund Commitments</div>
  <div class="tab"        onclick="showTab('pm',this)">People Moves</div>
</div>
<div id="em" class="panel active">
  <div class="filters">
    <label>Company</label><select class="co-filter" onchange="filterEM()"><option value="">All</option></select>
    <label>Scraped from</label><input type="date" onchange="filterEM()" style="width:140px;">
    <label>to</label><input type="date" onchange="filterEM()" style="width:140px;">
    <input type="text" placeholder="Search…" oninput="filterEM()" style="width:180px;">
  </div>
  <div class="count" id="em-count"></div>
  <table id="em-table"><thead><tr><th class="sortable" onclick="sortTable('em-table',0)">Stock Code</th><th>Firm ID</th><th class="sortable" onclick="sortTable('em-table',2)">Firm Name</th><th class="sortable" onclick="sortTable('em-table',3)">Firm Type</th><th class="sortable" onclick="sortTable('em-table',4)">BS Period</th><th>Telephone</th><th>Website</th><th>Address</th><th class="sortable" onclick="sortTable('em-table',8)">Scraped At</th><th>Action</th></tr></thead><tbody></tbody></table>
</div>
<div id="fs" class="panel">
  <div class="filters">
    <label>Company</label><select class="co-filter" onchange="filterFS()"><option value="">All</option></select>
    <input type="text" placeholder="Search…" oninput="filterFS()" style="width:180px;">
  </div>
  <div class="count" id="fs-count"></div>
  <table id="fs-table"><thead><tr><th class="sortable" onclick="sortTable('fs-table',0)">Stock Code</th><th>Firm ID</th><th class="sortable" onclick="sortTable('fs-table',2)">Firm Name</th><th>Firm Type</th><th class="sortable" onclick="sortTable('fs-table',4)">Period</th><th class="sortable" onclick="sortTable('fs-table',5)">Source</th><th class="sortable" onclick="sortTable('fs-table',6)">Total Assets (TWD, mn)</th><th class="sortable" onclick="sortTable('fs-table',7)">Inv. Property (TWD, mn)</th><th class="sortable" onclick="sortTable('fs-table',8)">Scraped At</th><th>Filing</th><th>Internal Notes</th><th>Action</th></tr></thead><tbody></tbody></table>
</div>
<div id="fc" class="panel">
  <div class="filters">
    <label>Company</label><select class="co-filter" onchange="filterFC()"><option value="">All</option></select>
    <label>Fund Type</label><select id="fc-ft-sel" onchange="filterFC()"><option value="">All</option></select>
    <label>From</label><input type="date" onchange="filterFC()" style="width:140px;">
    <label>To</label><input type="date" onchange="filterFC()" style="width:140px;">
    <input type="text" placeholder="Search fund name…" oninput="filterFC()" style="width:180px;">
  </div>
  <div class="count" id="fc-count"></div>
  <table id="fc-table"><thead><tr><th class="sortable" onclick="sortTable('fc-table',0)">Stock Code</th><th>Firm ID</th><th class="sortable" onclick="sortTable('fc-table',2)">Firm Name</th><th>Firm Type</th><th class="sortable" onclick="sortTable('fc-table',4)">Published Date</th><th class="sortable" onclick="sortTable('fc-table',5)">Fund Name</th><th class="sortable" onclick="sortTable('fc-table',6)">Fund Type</th><th class="sortable" onclick="sortTable('fc-table',7)">Commit Date</th><th class="sortable" onclick="sortTable('fc-table',8)">Amount</th><th>Headlines</th><th>Key Events</th><th class="sortable" onclick="sortTable('fc-table',11)">AUM as of</th><th class="sortable" onclick="sortTable('fc-table',12)">Scraped At</th><th>Internal Notes</th><th>Action</th></tr></thead><tbody></tbody></table>
</div>
<div id="pm" class="panel">
  <div class="filters">
    <label>Company</label><select class="co-filter" onchange="filterPM()"><option value="">All</option></select>
    <label>Status</label><select onchange="filterPM()"><option value="">All</option><option value="new">NEW</option><option value="historical">HISTORICAL</option><option value="changed">CHANGED</option></select>
    <label>From</label><input type="date" onchange="filterPM()" style="width:140px;">
    <label>To</label><input type="date" onchange="filterPM()" style="width:140px;">
    <input type="text" placeholder="Search name or role…" oninput="filterPM()" style="width:180px;">
  </div>
  <div class="count" id="pm-count"></div>
  <table id="pm-table"><thead><tr><th class="sortable" onclick="sortTable('pm-table',0)">Stock Code</th><th>Firm ID</th><th class="sortable" onclick="sortTable('pm-table',2)">Firm Name</th><th>Firm Type</th><th class="sortable" onclick="sortTable('pm-table',4)">Published Date</th><th class="sortable" onclick="sortTable('pm-table',5)">Role</th><th class="sortable" onclick="sortTable('pm-table',6)">New Holder</th><th class="sortable" onclick="sortTable('pm-table',7)">Previous Holder</th><th class="sortable" onclick="sortTable('pm-table',8)">Effective Date</th><th>Headlines</th><th>Key Events</th><th class="sortable" onclick="sortTable('pm-table',11)">AUM as of</th><th>Link</th><th class="sortable" onclick="sortTable('pm-table',13)">Scraped At</th><th>Internal Notes</th><th>Action</th></tr></thead><tbody></tbody></table>
</div>
""" + js + "\n</body></html>"

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    timestamped = OUTPUT_DIR / f"MOPSOV_{ts}.html"
    timestamped.write_text(html, encoding="utf-8")
    latest = OUTPUT_DIR / "report.html"
    latest.write_text(html, encoding="utf-8")
    logger.info("HTML report saved: %s (also → report.html)", timestamped.name)
    return latest


def write_excel(fund_commitments, people_moves, emops_data=None, balance_history=None, since=None, new_since=None):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    path = OUTPUT_DIR / f"MOPSOV_{ts}.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _link_font = Font(color="0563C1", underline="single", name="Calibri", size=10)

    # Summary
    ws = wb.create_sheet("Summary")
    new_label = f"New Fund Commitments (from {new_since})" if new_since else "New Fund Commitments"
    his_label = f"Historical Fund Commitments (before {new_since})" if new_since else "Historical Fund Commitments"
    _header(ws, ["Run Date", "Data Extracted From", new_label, his_label, "New People Moves", "Changed People Moves"])
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M"),
               since or "2 years",
               sum(1 for r in fund_commitments if r.get("status") == "NEW"),
               sum(1 for r in fund_commitments if r.get("status") == "HISTORICAL"),
               sum(1 for r in people_moves if r.get("status") == "NEW"),
               sum(1 for r in people_moves if r.get("status") == "CHANGED")])
    _autofit(ws)

    _co_map    = {w["stock_code"]: w["name_en"]          for w in WATCHLIST}
    _co_map.update({s["stock_code"]: s["company_name_en"] for s in _SUBSIDIARY_STUBS})
    _id_map    = {w["stock_code"]: w.get("firm_id", "")   for w in WATCHLIST}
    _type_map  = {w["stock_code"]: w.get("company_type", "") for w in WATCHLIST}
    _url_map   = {w["stock_code"]: w.get("firm_url", "")  for w in WATCHLIST}
    _firm_col  = 3   # Firm Name is column 3 in every sheet — carries the firm_url hyperlink

    # Fund Commitments — mirrors HTML FC table
    ws = wb.create_sheet("FundCommitments")
    _FC_COLS = ["Stock Code", "Firm ID", "Firm Name", "Firm Type", "Published Date", "Fund Name",
                "Fund Type", "Commit Date", "Amount", "Headlines", "Key Events", "AUM as of",
                "Scraped At", "Internal Notes"]
    _header(ws, _FC_COLS)
    _name_col  = _FC_COLS.index("Fund Name") + 1
    _amt_col   = _FC_COLS.index("Amount") + 1
    _fc_hl_col = _FC_COLS.index("Headlines") + 1
    for i, r in enumerate(fund_commitments, 2):
        code = r.get("stock_code", "")
        firm = _co_map.get(code, "")
        scraped = (r.get("scraped_at") or "")[:10].replace("-", "/")
        ws.append([code, _id_map.get(code, ""), firm, _type_map.get(code, ""),
                   r.get("announcement_date"), r.get("fund_name"),
                   r.get("fund_type"), r.get("commitment_date"), r.get("commitment_amount_raw"),
                   r.get("headline"), r.get("key_events"), r.get("bs_date"), scraped,
                   r.get("internal_notes", "")])
        ws.cell(i, _fc_hl_col).alignment = Alignment(wrap_text=True, vertical="top")
        firm_url = _url_map.get(code, "")
        if firm_url:
            cell = ws.cell(i, _firm_col); cell.hyperlink = firm_url; cell.font = _link_font
        if r.get("url"):
            cell = ws.cell(i, _name_col); cell.hyperlink = r["url"]; cell.font = _link_font
        if r.get("fx_url"):
            cell = ws.cell(i, _amt_col); cell.hyperlink = r["fx_url"]; cell.font = _link_font
        _status_fill(ws, i, r.get("status"))
    _autofit(ws)

    # People Moves — mirrors HTML PM table
    ws = wb.create_sheet("PeopleMoves")
    _PM_COLS = ["Stock Code", "Firm ID", "Firm Name", "Firm Type", "Published Date", "Role",
                "New Holder", "Previous Holder", "Effective Date", "Headlines", "Key Events",
                "AUM as of", "URL", "Status", "Scraped At", "Internal Notes"]
    _header(ws, _PM_COLS)
    _url_col   = _PM_COLS.index("URL") + 1
    _pm_hl_col = _PM_COLS.index("Headlines") + 1
    for i, r in enumerate(people_moves, 2):
        code = r.get("stock_code", "")
        role = r.get("role_title") or r.get("role_type") or ""
        firm = _co_map.get(code, "")
        scraped = (r.get("scraped_at") or "")[:10].replace("-", "/")
        ws.append([code, _id_map.get(code, ""), firm, _type_map.get(code, ""),
                   r.get("announcement_date"), role,
                   r.get("new_holder"), r.get("previous_holder"),
                   r.get("effective_date"), r.get("narrative_en"), r.get("key_events"),
                   r.get("bs_date", ""), r.get("url"), r.get("status"), scraped,
                   r.get("internal_notes", "")])
        ws.cell(i, _pm_hl_col).alignment = Alignment(wrap_text=True, vertical="top")
        firm_url = _url_map.get(code, "")
        if firm_url:
            cell = ws.cell(i, _firm_col); cell.hyperlink = firm_url; cell.font = _link_font
        if r.get("url"):
            cell = ws.cell(i, _url_col); cell.hyperlink = r["url"]; cell.font = _link_font
        _status_fill(ws, i, r.get("status"))
    _autofit(ws)

    # Company Profiles — Code, Name, Type, BS Period, Currency, Telephone, Website, Address
    if emops_data:
        ws = wb.create_sheet("CompanyProfiles")
        _EM_COLS = ["Stock Code", "Firm ID", "Firm Name", "Firm Type", "BS Period",
                    "Telephone", "Website", "Address", "Scraped At", "Notes"]
        _header(ws, _EM_COLS)
        _web_col   = _EM_COLS.index("Website") + 1
        _notes_col = _EM_COLS.index("Notes") + 1
        for i, r in enumerate(emops_data, 2):
            code = r.get("stock_code", "")
            if r.get("no_filing_data"):
                ws.append([code, _id_map.get(code, ""), r.get("company_name_en"),
                           _type_map.get(code, ""), "", "", "", "", "",
                           "Information not available on filings — please refer to firm website"])
                ws.cell(i, _notes_col).font = Font(italic=True, color="888888")
            else:
                scraped = (r.get("scraped_at") or "")[:10].replace("-", "/")
                ws.append([code, _id_map.get(code, ""),
                           r.get("name_en") or r.get("company_name_en"),
                           _type_map.get(code, ""), r.get("period"),
                           r.get("telephone"), r.get("web_address"),
                           r.get("address"), scraped, ""])
                web = r.get("web_address", "")
                if web:
                    url = web if web.startswith("http") else "https://" + web
                    cell = ws.cell(i, _web_col); cell.hyperlink = url; cell.font = _link_font
            firm_url = _url_map.get(code, "")
            if firm_url:
                cell = ws.cell(i, _firm_col); cell.hyperlink = firm_url; cell.font = _link_font
        _autofit(ws)

        # Financial Statements — annual (EMOPS) + quarterly (F26) balance sheet data
        ws = wb.create_sheet("FinancialStatements")
        _FS_COLS = ["Stock Code", "Firm ID", "Firm Name", "Firm Type", "Period", "Source",
                    "Total Assets (TWD, mn)", "Inv. Property (TWD, mn)", "Scraped At", "Filing",
                    "Internal Notes"]
        _header(ws, _FS_COLS)
        _filing_col  = _FS_COLS.index("Filing") + 1
        _ta_col      = _FS_COLS.index("Total Assets (TWD, mn)") + 1
        _fs_note_col = _FS_COLS.index("Internal Notes") + 1
        _warn_fill   = PatternFill("solid", fgColor="FFF3E0")
        _warn_font   = Font(color="E65100", bold=True)
        for i, r in enumerate(balance_history or [], 2):
            code = r.get("stock_code", "")
            base_name = r.get("name_en") or _co_map.get(code, "")
            co_name = f"{base_name} (Delisted)" if r.get("delisted") else base_name
            ta_val = "⚠ Manual check needed" if r.get("extraction_failed") else r.get("total_assets_raw")
            src_label = ("" if r.get("stub_only")
                         else "Balance Sheet" if r.get("source") == "annual"
                         else "PDFs - Consolidated FS" if r.get("source") == "quarterly_consolidated"
                         else "PDFs - Unconsolidated FS")
            filing_val = "View" if (r.get("source") == "annual" and r.get("filing_url")) else r.get("pdf_filename", "")
            ws.append([code, _id_map.get(code, ""), co_name, _type_map.get(code, ""),
                       r.get("period"), src_label, ta_val,
                       r.get("investment_property_raw"),
                       (r.get("scraped_at") or "")[:10].replace("-", "/"),
                       filing_val, r.get("internal_notes", "")])
            if r.get("extraction_failed"):
                ws.cell(i, _ta_col).font = _warn_font; ws.cell(i, _ta_col).fill = _warn_fill
            if r.get("source") == "annual" and r.get("filing_url"):
                cell = ws.cell(i, _filing_col); cell.hyperlink = r["filing_url"]; cell.font = _link_font
            if r.get("internal_notes"):
                ws.cell(i, _fs_note_col).alignment = Alignment(wrap_text=True, vertical="top")
            firm_url = _url_map.get(code, "")
            if firm_url:
                cell = ws.cell(i, _firm_col); cell.hyperlink = firm_url; cell.font = _link_font
        _autofit(ws)

    wb.save(path)
    logger.info("Excel saved: %s", path)
    return path

def _header(ws, cols):
    for c, h in enumerate(cols, 1):
        cell = ws.cell(1, c, h)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _status_fill(ws, row, status):
    if status == "NEW":
        for cell in ws[row]: cell.fill = _NEW_FILL
    elif status == "CHANGED":
        for cell in ws[row]: cell.fill = _CHG_FILL

_WIDE_COLS = {"Headline", "Headlines", "Narrative", "Key Events", "Internal Notes"}

def _autofit(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        col_name = col[0].value or ""
        width = max((len(str(c.value or "")) for c in col), default=10)
        max_w = 80 if col_name in _WIDE_COLS else 50
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_w)

# ── Terminal Preview ──────────────────────────────────────────────────────────

def print_results(fund_commitments, people_moves):
    print("\n" + "═" * 80)
    print(f"  FUND COMMITMENTS ({len(fund_commitments)} records)")
    print("═" * 80)
    for r in fund_commitments:
        print(f"\n  [{r.get('stock_code')}] {r.get('fund_name')} [{r.get('status')}]")
        print(f"    Type     : {r.get('fund_type')}")
        print(f"    Date     : {r.get('commitment_date')}")
        print(f"    Amount   : {r.get('commitment_currency')} {r.get('commitment_amount_raw')}")

    print("\n" + "═" * 80)
    print(f"  PEOPLE MOVES ({len(people_moves)} records)")
    print("═" * 80)
    for r in people_moves:
        print(f"\n  [{r.get('stock_code')}] {r.get('role_type')} [{r.get('status')}]")
        print(f"    New      : {r.get('new_holder')}")
        print(f"    Previous : {r.get('previous_holder')}")
        print(f"    Effective: {r.get('effective_date')}")
        print(f"    Narrative: {r.get('narrative_en')}")
    print("\n" + "═" * 80 + "\n")

# ── Main ──────────────────────────────────────────────────────────────────────


_EMOPS_TRACK = ["company_name_en", "address", "telephone", "web_address",
                "period", "currency", "total_assets_raw", "inv_property_raw"]

def _load_all_balance_history() -> list[dict]:
    """Load cached quarterly balance history from state files (no network calls)."""
    all_records = []
    for entry in WATCHLIST:
        code = entry["stock_code"]
        path = STATE_DIR / f"{code}_balance_history.json"
        if path.exists():
            try:
                all_records.extend(json.loads(path.read_text(encoding="utf-8")))
            except Exception:
                pass
    all_records.sort(key=lambda r: (r["stock_code"], -r.get("roc_year", 0), -r.get("season", 0)))
    return all_records

def _load_emops_data() -> list[dict]:
    """Read latest profile + balance sheet archive per company, detect field-level changes."""
    results = []
    for entry in WATCHLIST:
        code = entry["stock_code"]
        profile_files = sorted(ARCHIVE_DIR.glob(f"{code}_profile_*.json"), reverse=True)
        bs_files      = sorted(ARCHIVE_DIR.glob(f"{code}_balance_sheet_*.json"), reverse=True)
        profile, bs = {}, {}
        if profile_files:
            d = _load_json(profile_files[0])
            if d and d.get("records"):
                profile = d["records"][0]
        if bs_files:
            d = _load_json(bs_files[0])
            if d and d.get("records"):
                bs = d["records"][0]

        record = {
            "stock_code":           code,
            "name_en":              entry["name_en"],
            "company_type":         entry.get("company_type", ""),
            "company_name_en":      _clean_company_name(profile.get("company_name_en", "")),
            "address":              _clean_address(profile.get("address", "")),
            "telephone":            _format_tw_phone(profile.get("telephone", "")),
            "web_address":          _clean_web_address(profile.get("web_address", "")),
            "period":               bs.get("period", ""),
            "currency":             bs.get("currency", "").replace(" (thousands)", ""),
            "total_assets_raw":     bs.get("total_assets_raw", ""),
            "total_assets_numeric":  bs.get("total_assets_numeric"),
            "inv_property_raw":      bs.get("investment_property_raw", ""),
            "inv_property_numeric":  bs.get("investment_property_numeric"),
            "profile_status":       "",
            "changed_fields":       [],
            "scraped_at":           profile.get("scraped_at", ""),
        }

        # Compare with stored state to detect field-level changes
        state_path = STATE_DIR / f"{code}_emops.json"
        prev = _load_json(state_path) or {}
        if not prev:
            record["profile_status"] = "NEW"
        else:
            changed = [f for f in _EMOPS_TRACK
                       if record.get(f) and record.get(f) != prev.get(f)]
            record["profile_status"] = "CHANGED" if changed else "UNCHANGED"
            record["changed_fields"] = changed

        # Persist current values as new baseline
        state_path.parent.mkdir(parents=True, exist_ok=True)
        state_path.write_text(
            json.dumps({f: record.get(f, "") for f in _EMOPS_TRACK},
                       indent=2, ensure_ascii=False),
            encoding="utf-8")

        results.append(record)
    return results


def _fmt_millions(num) -> str:
    if num is None:
        return ""
    return f"{int(num / 1000):,}"

_PERIOD_SEASON = {"03/31": 1, "06/30": 2, "09/30": 3, "12/31": 4}

def _date_to_period_str(date_str: str) -> str:
    """Convert date string to period label like 'Q1 2025'."""
    m = re.match(r"(\d{4})[/-](\d{2})", date_str or "")
    if not m:
        return date_str or ""
    year, month = int(m.group(1)), int(m.group(2))
    if month <= 3: q = "Q1"
    elif month <= 6: q = "Q2"
    elif month <= 9: q = "Q3"
    else: q = "Q4"
    return f"{q} {year}"

def _build_fc_key_event(stock_code: str, date: str, fund_name: str,
                        formatted_amount: str, fund_type: str) -> str:
    firm_name = _firm_display_name(stock_code)
    period = _date_to_period_str(date)
    amount_part = f"committed {formatted_amount} to" if formatted_amount else "committed to"
    type_part = f", a {fund_type.lower()} fund" if fund_type else ""
    return f"In {period}, {firm_name} {amount_part} {fund_name}{type_part}."

def _build_pm_key_event(stock_code: str, role: str, new_holder: str,
                        prev_holder: str, date: str, change_type: str) -> str:
    firm_name = _firm_display_name(stock_code)
    period = _date_to_period_str(date)
    has_new  = bool(new_holder  and new_holder.lower()  not in ("none", "nil", "n/a", "na", ""))
    has_prev = bool(prev_holder and prev_holder.lower() not in ("none", "nil", "n/a", "na", ""))
    ct = change_type or ""
    if has_new:
        return f"In {period}, {firm_name} hired {new_holder} as its new {role}."
    elif has_prev:
        if "retirement" in ct:
            return f"In {period}, {prev_holder} retired as {role} of {firm_name}."
        return f"In {period}, {prev_holder} resigned as {role} of {firm_name}."
    return f"In {period}, {firm_name} announced a change in its {role}."

_SEP = "=" * 55

def _build_fc_internal_notes(r: dict) -> str:
    headline   = r.get("headline", "")
    key_events = r.get("key_events", "")
    amount     = r.get("commitment_amount_raw", "")
    fund_type  = r.get("fund_type", "")
    commit_date = r.get("commitment_date", "") or r.get("announcement_date", "")
    url        = r.get("url", "")
    meta_line = " | ".join(p for p in [commit_date, amount, fund_type] if p)
    hl_part  = f"Headlines: {headline}"  if headline   else ""
    ke_part  = f"Key Events: {key_events}" if key_events else ""
    fc_line  = " | ".join(p for p in [meta_line, hl_part, ke_part] if p)
    src_line = "https://mopsov.twse.com.tw/mops/web/ezsearch"
    if url:
        src_line += f" | {url}"
    return (f"PARTIAL UPDATE- Fund Commitment\n{_SEP}\n"
            f"Summary: \n{_SEP}\n"
            f"Fund Commitment: {fc_line}\n{_SEP}\n"
            f"Source: {src_line}\n{_SEP}")

def _build_pm_internal_notes(r: dict) -> str:
    role            = r.get("role_title", "") or r.get("role_type", "")
    announcement_date = r.get("announcement_date", "")
    effective_date  = r.get("effective_date", "")
    headline        = r.get("narrative_en", "")
    key_events      = r.get("key_events", "")
    url             = r.get("url", "")
    meta_line = " | ".join(p for p in [announcement_date, role, effective_date] if p)
    hl_part  = f"Headlines: {headline}"    if headline   else ""
    ke_part  = f"Key Events: {key_events}" if key_events else ""
    pm_line  = " | ".join(p for p in [meta_line, hl_part, ke_part] if p)
    src_line = "https://mopsov.twse.com.tw/mops/web/ezsearch"
    if url:
        src_line += f" | {url}"
    return (f"PARTIAL UPDATE- People Moves\n{_SEP}\n"
            f"Summary: \n{_SEP}\n"
            f"People Moves: {pm_line}\n{_SEP}\n"
            f"Source: {src_line}\n{_SEP}")

def _build_fs_internal_notes(ta: str, inv: str, period: str, pdf_filename: str,
                              firm_name: str = "", stock_code: str = "", url: str = "") -> str:
    firm_line = " | ".join(p for p in [
        f"Firm name: {firm_name}" if firm_name else "",
        f"Stock code: {stock_code}" if stock_code else "",
    ] if p)
    aum_line = f"AUM: {ta} mn | Real Estate: {inv} mn based on {period} filings"
    doc_line = f"added {pdf_filename}" if pdf_filename else ""
    src_line = "https://mopsov.twse.com.tw/mops/web/ezsearch"
    if url:
        src_line += f" | {url}"
    return (f"PARTIAL UPDATE- AUM & AA\n{_SEP}\n"
            f"Summary: \n{_SEP}\n"
            f"{firm_line}\n"
            f"{aum_line}\n{_SEP}\n"
            f"Document: {doc_line}\n{_SEP}\n"
            f"Source: {src_line}\n{_SEP}")

def _fs_filing_url(stock_code: str, period: str) -> str:
    """Return the MOPS iXBRL financial statement URL for a given company and period."""
    if not period or len(period) < 10:
        return ""
    year   = period[:4]
    season = _PERIOD_SEASON.get(period[5:], 0)
    if not season:
        return ""
    return (f"https://mopsov.twse.com.tw/server-java/t164sb01"
            f"?step=1&CO_ID={stock_code}&SYEAR={year}&SSEASON={season}&REPORT_ID=C")

def _build_fs_data(emops_data: list[dict], balance_history: list[dict],
                   since: str | None = None, new_since: str | None = None) -> list[dict]:
    """Combine annual EMOPS rows (financial holdings only) + quarterly PDF rows for the FS tab.
    Financial holdings identified by 'Financial Hold' in name_en.
    All numeric values are expressed in TWD millions (source data is in NT$K)."""
    rows = []
    _name_map = {w["stock_code"]: w["name_en"] for w in WATCHLIST}
    since_norm     = since.replace("/", "-")[:10]     if since     else None
    new_since_norm = new_since.replace("/", "-")[:10] if new_since else None

    # Annual BS rows: financial holding companies only
    for r in emops_data:
        if not r.get("total_assets_numeric"):
            continue
        if not _HOLDING_RE.search(r.get("name_en") or ""):
            continue
        period = r.get("period", "")
        if since_norm and period:
            if period.replace("/", "-")[:10] < since_norm:
                continue
        entry = _WATCHLIST_MAP.get(r["stock_code"], {})
        is_new = bool(new_since_norm and period and period.replace("/", "-")[:10] >= new_since_norm)
        ta  = _fmt_millions(r.get("total_assets_numeric")) or "—"
        inv = _fmt_millions(r.get("inv_property_numeric")) or "—"
        internal_notes = _build_fs_internal_notes(
            ta, inv, period, "",
            firm_name=r["name_en"], stock_code=r["stock_code"],
            url=_fs_filing_url(r["stock_code"], period),
        ) if is_new else ""
        rows.append({
            "stock_code":              r["stock_code"],
            "name_en":                 r["name_en"],
            "period":                  period,
            "currency":                "TWD",
            "total_assets_raw":        _fmt_millions(r.get("total_assets_numeric")),
            "investment_property_raw": _fmt_millions(r.get("inv_property_numeric")),
            "scraped_at":              (r.get("scraped_at") or "")[:10],
            "delisted":                entry.get("delisted", False),
            "source":                  "annual",
            "filing_url":              _fs_filing_url(r["stock_code"], period),
            "pdf_path":                "",
            "pdf_filename":            "",
            "internal_notes":          internal_notes,
        })

    # All quarterly rows from cached PDF-extracted balance history.
    # Only show rows where a PDF was successfully downloaded — any season without a local pdf_path is skipped.
    _stub_name_map = {s["stock_code"]: s["company_name_en"] for s in _SUBSIDIARY_STUBS}
    for r in balance_history:
        season = r.get("season", 4)
        if not r.get("pdf_path") and r.get("source") != "mops_ixbrl":
            continue
        entry  = _WATCHLIST_MAP.get(r["stock_code"], {})
        yr     = r.get("roc_year", 0)
        period_end = (f"{yr + 1911}/{_FS_SEASON_DATE.get(season, '12/31')}"
                      if yr else r.get("period", ""))

        if since_norm and period_end:
            if period_end.replace("/", "-")[:10] < since_norm:
                continue

        internal_sub_code = str(r.get("subsidiary_code") or "")
        display_code = (_FS_SUBSIDIARY_CODE_MAP.get(internal_sub_code)
                        or r["stock_code"])
        sub_name = (_WATCHLIST_MAP.get(display_code, {}).get("name_en")
                    or _stub_name_map.get(display_code)
                    or r.get("subsidiary_name_en")
                    or entry.get("f26_name_en")
                    or _name_map.get(display_code, display_code))
        source = (r.get("source") if r.get("source") == "mops_ixbrl"
                  else "quarterly_consolidated" if r.get("is_consolidated") else "quarterly")
        is_new = bool(new_since_norm and period_end and period_end.replace("/", "-")[:10] >= new_since_norm)
        pdf_filename = Path(r.get("pdf_path", "")).name if r.get("pdf_path") else ""
        ta  = _fmt_millions(r.get("total_assets_numeric")) or "—"
        inv = _fmt_millions(r.get("investment_property_numeric")) or "—"
        internal_notes = _build_fs_internal_notes(
            ta, inv, period_end, pdf_filename,
            firm_name=sub_name, stock_code=display_code,
            url="" if pdf_filename else _fs_filing_url(display_code, period_end),
        ) if is_new else ""

        rows.append({
            "stock_code":              display_code,
            "name_en":                 sub_name,
            "period":                  period_end,
            "currency":                "TWD",
            "total_assets_raw":        _fmt_millions(r.get("total_assets_numeric")),
            "investment_property_raw": _fmt_millions(r.get("investment_property_numeric")),
            "scraped_at":              (r.get("scraped_at") or "")[:10],
            "delisted":                entry.get("delisted", False),
            "source":                  source,
            "filing_url":              r.get("filing_url", "") if source == "mops_ixbrl" else "",
            "pdf_path":                r.get("pdf_path", ""),
            "pdf_filename":            pdf_filename,
            "extraction_failed":       r.get("extraction_failed", False),
            "internal_notes":          internal_notes,
        })

    # Deduplicate quarterly rows (duplicate state entries can arise when the same PDF is stored
    # under two different keys). Keep the first record that has total_assets_numeric; else first seen.
    seen_qtr: dict[tuple, dict] = {}
    deduped = []
    for r in rows:
        if r.get("source", "").startswith("quarterly") or r.get("source") == "mops_ixbrl":
            key = (r["stock_code"], r["period"])
            if key not in seen_qtr:
                seen_qtr[key] = r
                deduped.append(r)
            elif r.get("total_assets_numeric") and not seen_qtr[key].get("total_assets_numeric"):
                # Replace with the record that actually has data
                deduped[deduped.index(seen_qtr[key])] = r
                seen_qtr[key] = r
        else:
            deduped.append(r)
    rows = deduped

    # Deduplicate: if annual and quarterly share the same (stock_code, period, name_en), keep annual.
    # name_en is included so subsidiary rows (e.g. Fubon Life under 2881) are not dropped.
    annual_keys = {(r["stock_code"], r["period"], r["name_en"]) for r in rows if r["source"] == "annual"}
    rows = [r for r in rows if r["source"] == "annual"
            or (r["stock_code"], r["period"], r["name_en"]) not in annual_keys]

    # Add stub rows for ALL WATCHLIST + _SUBSIDIARY_STUBS entries that have no FS data at all,
    # so every tracked company always appears in the FS tab.
    present_codes = {r["stock_code"] for r in rows}
    all_tracked = (
        [(w["stock_code"], w["name_en"], w.get("delisted", False)) for w in WATCHLIST]
        + [(s["stock_code"], s["company_name_en"], False) for s in _SUBSIDIARY_STUBS]
    )
    for code, name, delisted in all_tracked:
        if code not in present_codes:
            rows.append({
                "stock_code":              code,
                "name_en":                 name,
                "period":                  "",
                "currency":                "TWD",
                "total_assets_raw":        "",
                "investment_property_raw": "",
                "scraped_at":              "",
                "delisted":                delisted,
                "source":                  "",
                "filing_url":              "",
                "pdf_path":                "",
                "pdf_filename":            "",
                "extraction_failed":       False,
                "internal_notes":          "",
                "stub_only":               True,
            })

    return rows


async def run(companies=None, export_excel=True, mode="full", since=None, new_since=None):
    """
    mode="full"       — scrape everything: EMOPS profiles, quarterly PDF balance sheets, FC, PM (run quarterly)
    mode="daily"      — FC + PM only; EMOPS/FS loaded from cached state files (run nightly)
    mode="report-only" — no network calls; rebuild HTML/Excel from cached state files only
    """
    watchlist = WATCHLIST if not companies else [w for w in WATCHLIST if w["stock_code"] in companies]
    scrape_profiles = mode == "full"
    scrape_fc_pm    = mode != "report-only"
    logger.info("Running MOPSOV [mode=%s] for %d companies", mode, len(watchlist))
    if since:
        logger.info("Extracting fund commitments from %s", since)
    if new_since:
        logger.info("Last-run baseline date: %s", new_since)

    all_funds, all_people, all_balance_history = [], [], []

    for entry in watchlist:
        code = entry["stock_code"]
        logger.info("── %s %s", code, entry["name_en"])

        if scrape_profiles:
            profile = await scrape_emops_profile(code)
            profile = detect_changes([profile], STATE_DIR / f"{code}_profile.json", ["stock_code"])[0]
            archive(code, "profile", [profile])

            bs = await scrape_emops_balance_sheet(code)
            bs = detect_changes([bs], STATE_DIR / f"{code}_balance_sheet.json", ["stock_code"])[0]
            archive(code, "balance_sheet", [bs])
            logger.info("Profile [%s] addr=%s | BS period=%s total_assets=%s",
                        code, profile.get("address") or "—", bs.get("period") or "—",
                        bs.get("total_assets_raw") or "—")

            subs = entry.get("f26_subsidiaries") or None
            bs_history = await scrape_quarterly_reports(
                code,
                subsidiary_name_en=entry.get("f26_name_en", ""),
                subsidiaries=subs,
                use_consolidated=bool(_HOLDING_RE.search(entry.get("name_en", ""))) and not subs,
            )
            all_balance_history.extend(bs_history)

        if scrape_fc_pm:
            funds = await scrape_fund_commitments(code, sdate=since)
            funds = detect_changes(funds, STATE_DIR / f"{code}_funds.json",
                                   ["stock_code", "fund_name", "commitment_date"])
            funds = _apply_date_status(funds, new_since, "announcement_date")
            all_funds.extend(funds)
            archive(code, "fund_commitments", funds)

            moves = await scrape_people_moves(code, sdate=since)
            moves = detect_changes(moves, STATE_DIR / f"{code}_people.json",
                                   ["stock_code", "role_type", "change_date"])
            moves = _apply_date_status(moves, new_since, "announcement_date")
            all_people.extend(moves)
            archive(code, "people_moves", moves)

    # Always build FS from state files so partial runs (--companies X) still show all companies
    all_balance_history = _load_all_balance_history()
    _populate_aum_cache(all_balance_history)
    # Note: period filtering is now handled inside _build_fs_data using period end dates

    # In report-only mode, load FC and PM from latest archive files instead of scraping
    if mode == "report-only":
        all_funds, all_people = [], []
        for entry in WATCHLIST:
            code = entry["stock_code"]
            for arc in sorted(ARCHIVE_DIR.glob(f"{code}_fund_commitments_*.json"), reverse=True)[:1]:
                d = _load_json(arc)
                if d and d.get("records"):
                    all_funds.extend(d["records"])
            for arc in sorted(ARCHIVE_DIR.glob(f"{code}_people_moves_*.json"), reverse=True)[:1]:
                d = _load_json(arc)
                if d and d.get("records"):
                    for r in d["records"]:
                        # Re-resolve subsidiary code and re-clean names/narrative
                        r["stock_code"] = _resolve_subsidiary(r["stock_code"], r.get("subject", ""))
                        r["new_holder"] = _clean_holder_name(r.get("new_holder", ""))
                        r["previous_holder"] = _clean_holder_name(r.get("previous_holder", ""))
                        r["narrative_en"] = _NARRATIVE_ABBREV_RE.sub(
                            "", _build_narrative(
                                r["stock_code"], r.get("role_title", ""),
                                r["new_holder"], r["previous_holder"],
                                r.get("change_type", ""), r.get("effective_date", ""),
                            )
                        ).strip()
                        r["key_events"] = _build_pm_key_event(
                            r["stock_code"], r.get("role_title", ""),
                            r["new_holder"], r["previous_holder"],
                            r.get("effective_date", "") or r.get("change_date", "") or r.get("announcement_date", ""),
                            r.get("change_type", ""),
                        )
                        _, r["bs_date"] = _get_latest_aum(r["stock_code"])
                    all_people.extend(d["records"])

        # Re-compute fund_type, subsidiary codes, formatted amounts, headlines, and key_events for FC
        for r in all_funds:
            r["stock_code"] = _fc_resolve_code(r["stock_code"], r.get("subject", ""))
            r["fund_name"] = _clean_fund_name(r.get("fund_name", ""))
            r["fund_type"] = _normalize_fund_type(r.get("fund_type", ""), r.get("fund_name", ""))
            formatted, fx = await _format_commitment_amount(
                r.get("commitment_amount_raw", ""), r.get("commitment_currency", "")
            )
            amt = formatted or r.get("commitment_amount_raw", "")
            if formatted:
                twd_m = re.search(r"TWD ([\d,]+) million", formatted)
                r["twd_amount_mn"] = twd_m.group(1) if twd_m else r.get("twd_amount_mn", "")
                r["fx_url"] = fx
            if r.get("fund_type"):
                r["headline"] = _build_fund_headline(
                    r["stock_code"], r.get("fund_name", ""), amt, r.get("fund_type", ""))
                r["key_events"] = _build_fc_key_event(
                    r["stock_code"],
                    r.get("commitment_date", "") or r.get("announcement_date", ""),
                    r.get("fund_name", ""), amt, r.get("fund_type", ""))

    # Re-apply date-based status so internal_notes fire for every record inside
    # the --new-since window, regardless of mode or whether the record existed in
    # a prior run.  In report-only mode _apply_date_status was never called, so
    # this is the primary gate.  In full mode it is a harmless re-application.
    all_funds  = _apply_date_status(all_funds,  new_since, "announcement_date")
    all_people = _apply_date_status(all_people, new_since, "announcement_date")

    # Generate internal_notes for every NEW record (always regenerate, never rely
    # on a cached value from a prior archive so the text reflects the latest AUM).
    for r in all_funds:
        if r.get("status") == "NEW":
            r["internal_notes"] = _build_fc_internal_notes(r)
        else:
            r["internal_notes"] = ""

    for r in all_people:
        if r.get("status") == "NEW":
            r["internal_notes"] = _build_pm_internal_notes(r)
        else:
            r["internal_notes"] = ""

    print_results(all_funds, all_people)
    if export_excel:
        emops_data = _load_emops_data()
        emops_data = emops_data + _SUBSIDIARY_STUBS
        fs_data = _build_fs_data(emops_data, all_balance_history, since=since, new_since=new_since)

        # Warn for WATCHLIST companies with no FS data at all
        fs_codes = {r["stock_code"] for r in fs_data}
        for entry in WATCHLIST:
            code = entry["stock_code"]
            if not entry.get("delisted"):
                has_fs = any(r.startswith(code) for r in fs_codes)
                if not has_fs:
                    logger.warning("No FS data found for %s (%s)", code, entry["name_en"])

        write_excel(all_funds, all_people, emops_data, balance_history=fs_data, since=since, new_since=new_since)
        write_html_report(all_funds, all_people, emops_data, balance_history=fs_data, since=since, new_since=new_since)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="MOPSOV Scraper — Fund Commitments & People Moves")
    parser.add_argument("--companies", nargs="+", help="Limit to specific stock codes e.g. 2882 2330")
    parser.add_argument("--no-excel", action="store_true", help="Print only, skip Excel/HTML export")
    parser.add_argument("--mode", choices=["full", "daily", "report-only"], default="full",
                        help="full=everything (default); daily=FC+PM only; report-only=rebuild output from cache, no network")
    parser.add_argument("--since", help="Extract data from this date e.g. 2025/01/01 (YYYY/MM/DD)")
    parser.add_argument("--new-since", dest="new_since", help="Flag as NEW if on/after this date e.g. 2025/09/30")
    args = parser.parse_args()

    asyncio.run(run(companies=args.companies, export_excel=not args.no_excel,
                    mode=args.mode, since=args.since, new_since=args.new_since))
